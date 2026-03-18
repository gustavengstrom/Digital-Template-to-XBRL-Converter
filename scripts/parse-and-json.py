"""Extract facts from an Excel digital template and output a JSON mapping of labels to values."""

import argparse
import json
import logging
import re
from contextlib import closing
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import rich.traceback
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.formula import ArrayFormula
from rich.logging import RichHandler

import mireport
from mireport.conversionresults import ConversionResultsBuilder
from mireport.excelprocessor import VSME_DEFAULTS, ExcelProcessor
from mireport.excelutil import (
    CellValueType,
    checkExcelFilePath,
    getNamedRanges,
    loadExcelFromPathOrFileLike,
)
from mireport.localise import EU_LOCALES, argparse_locale


def createArgParser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract facts from Excel digital template and output as JSON."
    )
    parser.add_argument("excel_file", type=Path, help="Path to the Excel file")
    parser.add_argument(
        "output_path",
        type=Path,
        help="Path to save the JSON output file (e.g. output/report.json)",
    )
    parser.add_argument(
        "--output-locale",
        type=argparse_locale,
        default=None,
        help=f"Locale to use when formatting values (default: None). Examples:\n{sorted(EU_LOCALES)}",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Suppress overwrite warnings and force file replacement.",
    )
    parser.add_argument(
        "--flat",
        action="store_true",
        default=False,
        help="Output a flat label→value mapping instead of the detailed structure.",
    )
    return parser


def parseArgs(parser: argparse.ArgumentParser) -> argparse.Namespace:
    args = parser.parse_args()
    return args


def fact_value_to_json(value) -> str | int | float | bool | None:
    """Convert a FactValue to a JSON-serializable type."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


# ---------------------------------------------------------------------------
# JsonExcelProcessor – subclass that also captures template_* named ranges
# ---------------------------------------------------------------------------


class JsonExcelProcessor(ExcelProcessor):
    """
    Extends ExcelProcessor to additionally capture the template_* named ranges
    that the base class intentionally skips (they are not XBRL facts but carry
    section-applicability flags and human-readable labels).

    The base class closes the workbook at the end of ``populateReport()``.
    This subclass overrides that method to also call ``get_template_data()``
    while the workbook is still open, storing the result for later retrieval.
    """

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # {name -> DefinedName} for every template_* range in the workbook
        self._template_defined_names: dict[str, DefinedName] = {}
        self._template_data: Optional[dict] = None

    # -- overrides -----------------------------------------------------------

    def _recordNamedRanges(self) -> None:
        """Record both XBRL-related *and* template_* named ranges."""
        # Let the base class populate _unusedDefinedNames (non-template ranges)
        super()._recordNamedRanges()

        # Additionally collect every template_* range that the base class skips
        for dn in self._workbook.defined_names.values():
            if dn.name and dn.name.startswith("template_"):
                self._template_defined_names[dn.name] = dn

    def populateReport(self):
        """
        Override to extract template data while the workbook is still open.
        The base class closes the workbook in its ``finally`` block, so we
        hook into ``checkForUnhandledItems`` which runs just before that.
        """
        from mireport.xbrlreport import InlineReport

        report = super().populateReport()
        return report

    def checkForUnhandledItems(self) -> None:
        """
        Called by the base class just before the workbook is closed.
        We extract template data here while the workbook is still available,
        then call through to the base implementation.
        """
        self._template_data = self._extract_template_data()
        super().checkForUnhandledItems()

    def get_template_data(self) -> dict:
        """
        Return the template data collected during ``populateReport()``.
        Must be called after ``populateReport()`` has completed.
        """
        if self._template_data is None:
            raise RuntimeError(
                "Template data not yet available. Call populateReport() first."
            )
        return self._template_data

    def _extract_template_data(self) -> dict:
        """
        Internal: extract template data while the workbook is still open.

        Return a structured dict with four top-level keys:

        * ``labels``   – template_label_* display strings
        * ``metadata`` – template_reporting_* and other template_* descriptors
        * ``sectionApplicability`` – per-section applicability status parsed
          from the descriptors ("always", "conditional", "optional", …)
        * ``omittedDisclosures`` – list of disclosure section names the user
          has marked as omitted via the List of Omitted Disclosures dropdown
          (from the ``ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitive
          Information`` named range)
        * ``enumLists`` – all ``enum_*`` defined names from the workbook mapped
          to their list of allowed values (the dropdown option lists)
        * ``xbrlToEnum`` – mapping from XBRL named-range names to the
          ``enum_*`` list name(s) that provide their dropdown options (derived
          from the Excel data-validation rules on the corresponding cells)
        """
        assert self._workbook is not None
        all_ranges, _ = getNamedRanges(self._workbook)

        labels: dict[str, object] = {}
        metadata: dict[str, object] = {}
        section_applicability: dict[str, str] = {}

        # --- labels & metadata ------------------------------------------------
        for name, cells in sorted(all_ranges.items()):
            if not name.startswith("template_"):
                continue
            non_empty = [c for c in cells if c is not None]
            if not non_empty:
                continue
            values = [fact_value_to_json(v) for v in non_empty]
            value = values[0] if len(values) == 1 else values

            if name.startswith("template_label_"):
                clean = name[len("template_label_") :]
                labels[clean] = value
            else:
                clean = name[len("template_") :]
                metadata[clean] = value

        # --- i18n labels from the Translations sheet --------------------------
        # The Translations sheet has: col A = label ID, col B = active language,
        # cols C-K = individual language translations (en, da, fr, de, it, lt,
        # pl, pt, es).  We build a dict: label_key -> {lang_code: text}.
        translations: dict[str, dict[str, str]] = {}
        if "Translations" in self._workbook:
            tws = self._workbook["Translations"]
            # Row 2 has language codes
            lang_codes: dict[int, str] = {}
            for col in range(3, tws.max_column + 1):
                code = tws.cell(row=2, column=col).value
                if code:
                    lang_codes[col] = str(code).strip().strip("'")
            # Rows 4+ have labels
            for row in range(4, tws.max_row + 1):
                label_id = tws.cell(row=row, column=1).value
                if not label_id:
                    continue
                label_id = str(label_id).strip().strip("'")
                if not label_id.startswith("template_label_"):
                    continue
                clean = label_id[len("template_label_") :]
                lang_map: dict[str, str] = {}
                for col, lang in lang_codes.items():
                    v = tws.cell(row=row, column=col).value
                    if v:
                        lang_map[lang] = str(v).strip("'")
                if lang_map:
                    translations[clean] = lang_map

        # --- section applicability -------------------------------------------
        for key, val in metadata.items():
            text = str(val)
            if "[Always to be reported + If applicable]" in text:
                status = "always+conditional"
            elif "[Always to be reported]" in text:
                status = "always"
            elif "[If applicable linked with" in text:
                status = "conditional-linked"
            elif "[If applicable]" in text:
                status = "conditional"
            elif "[May (optional)]" in text:
                status = "optional"
            else:
                continue
            section_applicability[key] = status

        # --- omitted disclosures -------------------------------------------
        # The user can explicitly mark whole disclosure sections as omitted
        # via a multi-select dropdown.  The values come from the XBRL named
        # range ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitiveInformation
        # which maps to a column in General Information (E123:E222).
        omitted_disclosures: list[str] = []
        omitted_range_name = (
            "ListOfOmittedDisclosuresDeemedToBeClassifiedOrSensitiveInformation"
        )
        if omitted_range_name in all_ranges:
            for val in all_ranges[omitted_range_name]:
                if val is not None:
                    s = str(val).strip()
                    if s and s.lower() != "none":
                        omitted_disclosures.append(s)
        else:
            # Fall back: try reading directly from the workbook defined name
            for dn in self._workbook.defined_names.values():
                if dn.name == omitted_range_name:
                    for sheet_title, coord in dn.destinations:
                        if sheet_title in self._workbook:
                            ws = self._workbook[sheet_title]
                            try:
                                cr = CellRange(coord)
                            except Exception:
                                break
                            for row in range(cr.min_row, cr.max_row + 1):
                                for col in range(cr.min_col, cr.max_col + 1):
                                    cv = ws.cell(row=row, column=col).value
                                    if cv is not None:
                                        s = str(cv).strip()
                                        if s and s.lower() != "none":
                                            omitted_disclosures.append(s)

        # --- all defined names ------------------------------------------------
        # Build a complete catalogue of every defined name in the workbook
        # together with its cell reference and current value(s).  This ensures
        # no defined name is silently dropped from the JSON output.
        defined_names: dict[str, dict] = {}
        for dn in sorted(
            self._workbook.defined_names.values(), key=lambda d: d.name or ""
        ):
            if not dn.name:
                continue
            # Determine category
            if dn.name.startswith("template_label_"):
                cat = "template_label"
            elif dn.name.startswith("template_"):
                cat = "template"
            elif dn.name.startswith("enum_"):
                cat = "enum"
            else:
                cat = "xbrl"

            # Read value(s) from the cell range
            values: list = []
            cell_ref = dn.attr_text  # e.g. "'General Information'!$E$123:$E$222"
            for sheet_title, coord in dn.destinations:
                if sheet_title not in self._workbook:
                    continue
                ws = self._workbook[sheet_title]
                try:
                    cr = CellRange(coord)
                except Exception:
                    continue
                for row in range(cr.min_row, cr.max_row + 1):
                    for col in range(cr.min_col, cr.max_col + 1):
                        cv = ws.cell(row=row, column=col).value
                        if cv is not None:
                            values.append(fact_value_to_json(cv))

            entry: dict = {
                "category": cat,
                "cellRef": cell_ref,
            }
            if values:
                entry["value"] = values[0] if len(values) == 1 else values
            else:
                entry["value"] = None

            defined_names[dn.name] = entry

        # --- enum lists -------------------------------------------------------
        # Extract all enum_* defined names and their values.  These are the
        # dropdown option lists used by the template.
        enum_lists: dict[str, list[str]] = {}
        for dn in self._workbook.defined_names.values():
            if not dn.name or not dn.name.startswith("enum_"):
                continue
            values: list[str] = []
            for sheet_title, coord in dn.destinations:
                if sheet_title not in self._workbook:
                    continue
                ws = self._workbook[sheet_title]
                try:
                    cr = CellRange(coord)
                except Exception:
                    continue
                for row in range(cr.min_row, cr.max_row + 1):
                    for col in range(cr.min_col, cr.max_col + 1):
                        cv = ws.cell(row=row, column=col).value
                        if cv is not None:
                            values.append(str(cv))
            enum_lists[dn.name] = values

        # --- XBRL → enum mapping via data validation -------------------------
        # Cells with dropdowns have data-validation rules whose formula1
        # references an enum_* defined name.  By intersecting those cells with
        # the cells covered by XBRL named ranges we can link each XBRL concept
        # to the enum list(s) that provide its answer options.
        data_sheet_names = [
            "General Information",
            "Environmental Disclosures",
            "Social Disclosures",
            "Governance Disclosures",
        ]
        # (sheet, row, col) → enum_name
        cell_to_enum: dict[tuple[str, int, int], str] = {}
        for sheet_name in data_sheet_names:
            if sheet_name not in self._workbook:
                continue
            ws = self._workbook[sheet_name]
            if ws.data_validations is None:
                continue
            for dv in ws.data_validations.dataValidation:
                formula = dv.formula1
                if not formula or not formula.startswith("enum_"):
                    continue
                enum_name = formula.strip()
                for cr_str in str(dv.sqref).split():
                    try:
                        cr = CellRange(cr_str)
                    except Exception:
                        continue
                    for row in range(cr.min_row, cr.max_row + 1):
                        for col in range(cr.min_col, cr.max_col + 1):
                            cell_to_enum[(sheet_name, row, col)] = enum_name

        xbrl_to_enum: dict[str, list[str]] = {}
        for dn in self._workbook.defined_names.values():
            if not dn.name or dn.name.startswith(("enum_", "template_")):
                continue
            for sheet_title, coord in dn.destinations:
                if sheet_title not in self._workbook:
                    continue
                try:
                    cr = CellRange(coord)
                except Exception:
                    continue
                for row in range(cr.min_row, cr.max_row + 1):
                    for col in range(cr.min_col, cr.max_col + 1):
                        key = (sheet_title, row, col)
                        if key in cell_to_enum:
                            enum_name = cell_to_enum[key]
                            xbrl_to_enum.setdefault(dn.name, [])
                            if enum_name not in xbrl_to_enum[dn.name]:
                                xbrl_to_enum[dn.name].append(enum_name)

        # --- validation rules (from IF-formula cells) -------------------------
        # Data sheets contain cells whose formulas reference
        # ``template_label_ok`` and ``template_label_missing_value`` to
        # signal whether an input cell has been filled.  By parsing these
        # formulas we can determine:
        #   1. Which XBRL fields are validated (have a validation cell)
        #   2. Whether the field is *required* (formula can emit
        #      MISSING VALUE) or merely *informational*
        #   3. Whether the validation is *conditional* on another cell
        #      (e.g. a boolean question or the BasisForPreparation choice)
        #
        # The main workbook is loaded with ``data_only=True`` so formulas are
        # replaced by cached values.  We load a second copy with
        # ``data_only=False`` to read the raw IF formulas.
        #
        # The result is ``validationRules``: a dict keyed by XBRL field name
        # with ``{required, conditionField, validationCell, status}``.

        from openpyxl import load_workbook as _load_wb

        validation_rules: dict[str, dict] = {}
        section_labels: dict[str, str] = {}

        try:
            wb_formulas = _load_wb(
                filename=self._excelPathOrFileLike,
                read_only=False,
                data_only=False,
            )
        except Exception:
            wb_formulas = None

        if wb_formulas is not None:
            # Build (sheet, row, col) → XBRL name for the entire workbook
            _xbrl_cell_map: dict[tuple[str, int, int], str] = {}
            for dn in wb_formulas.defined_names.values():
                if not dn.name or dn.name.startswith(("template_", "enum_")):
                    continue
                for _sh, _co in dn.destinations:
                    if _sh not in wb_formulas:
                        continue
                    try:
                        _cr = CellRange(_co)
                    except Exception:
                        continue
                    for _r in range(_cr.min_row, _cr.max_row + 1):
                        for _c in range(_cr.min_col, _cr.max_col + 1):
                            _xbrl_cell_map[(_sh, _r, _c)] = dn.name

            # Helper: convert column letters to 1-based column index
            def _col_to_idx(col_letters: str) -> int:
                idx = 0
                for i, ch in enumerate(reversed(col_letters)):
                    idx += (ord(ch) - ord("A") + 1) * (26**i)
                return idx

            # Set of XBRL defined names that can appear verbatim in formulas
            # (e.g. BasisForPreparation, BasisForReporting).  Used for fast
            # named-range condition detection (step a).
            _xbrl_defined_names: set[str] = {
                dn.name
                for dn in wb_formulas.defined_names.values()
                if dn.name and not dn.name.startswith(("template_", "enum_"))
            }

            # --- Build yellow cell + row→label maps for condition resolution --
            # Yellow cells (fill FFFFFF99) serve as condition triggers (e.g.
            # boolean questions) in validation formulas.  We build:
            #   _yellow_cell_set: positions of yellow cells
            #   _vr_row_to_label: (sheet, row) → template_label key
            #   _vr_col_label: (sheet, row, col) → template_label key
            #     for column-aware upward scan in yellow-cell resolution.
            #   _vr_cell_id_map: (sheet, row, col) → identifier
            #     that resolves ANY cell ref to an XBRL name, yellow-cell
            #     template_label key, or None.

            def _is_input_cell_fill(cell) -> bool:
                """Return True if the cell has an input-cell fill colour.

                Two fill patterns denote user-editable input cells in the
                VSME template:
                  1. Explicit yellow: RGB ``FFFFFF99``
                  2. Light-gray theme: theme colour 0 with tint ≈ −0.15
                     (used for numeric/boolean inputs that feed into
                     calculated XBRL fields such as board-member counts
                     and employee turnover figures).
                """
                if not (cell.fill and cell.fill.start_color):
                    return False
                sc = cell.fill.start_color
                if sc.index == "FFFFFF99":
                    return True
                if (
                    sc.type == "theme"
                    and sc.index == 0
                    and sc.tint is not None
                    and abs(sc.tint + 0.1499984740745262) < 0.001
                ):
                    return True
                return False

            def _is_computed_cell_fill(cell) -> bool:
                """Return True if the cell has a computed-cell fill colour.

                Cells with theme colour 8 and tint ≈ 0.80 are
                auto-calculated formula cells that require no user
                input (e.g. total hours worked, accident rates,
                turnover rates, gender ratios).
                """
                if not (cell.fill and cell.fill.start_color):
                    return False
                sc = cell.fill.start_color
                if (
                    sc.type == "theme"
                    and sc.tint is not None
                    and abs(sc.tint - 0.7999816888943144) < 0.001
                ):
                    return True
                return False

            _vr_label_re = re.compile(r"^=?(template_label_\w+)")
            _yellow_cell_set: set[tuple[str, int, int]] = set()
            _computed_cell_set: set[tuple[str, int, int]] = set()
            _vr_row_to_label: dict[tuple[str, int], str] = {}
            _vr_col_label: dict[tuple[str, int, int], str] = {}

            for _sn in data_sheet_names:
                if _sn not in self._workbook or _sn not in wb_formulas:
                    continue
                _ws_v = self._workbook[_sn]
                _ws_f = wb_formulas[_sn]
                # Collect input cell positions from the value workbook
                # (yellow fill FFFFFF99 or light-gray theme:0/tint:-0.15)
                # Also collect computed cells (theme tint ≈ 0.80).
                for _row_cells in _ws_v.iter_rows(
                    min_row=1,
                    max_row=_ws_v.max_row,
                    min_col=1,
                    max_col=_ws_v.max_column,
                ):
                    for _c in _row_cells:
                        if _is_input_cell_fill(_c):
                            _yellow_cell_set.add((_sn, _c.row, _c.column))
                        elif _is_computed_cell_fill(_c):
                            _computed_cell_set.add((_sn, _c.row, _c.column))
                # Build (sheet, row) → leftmost template_label key from formulas
                # and (sheet, row, col) → template_label key for column-aware resolution
                for _row_cells in _ws_f.iter_rows(
                    min_row=1,
                    max_row=_ws_f.max_row,
                    min_col=1,
                    max_col=_ws_f.max_column,
                ):
                    for _c in _row_cells:
                        if isinstance(_c.value, str):
                            _m = _vr_label_re.match(_c.value.strip())
                            if _m and "warning" not in _m.group(1):
                                _k = (_sn, _c.row)
                                if _k not in _vr_row_to_label:
                                    _vr_row_to_label[_k] = _m.group(1)
                                _vr_col_label[(_sn, _c.row, _c.column)] = _m.group(1)

            def _resolve_cell_id(
                sheet: str, row: int, col: int
            ) -> tuple[str | None, str]:
                """Resolve a cell reference to (identifier, category).

                Categories:
                  "xbrl"    – XBRL defined name
                  "yellow"  – yellow cell with template_label key
                  "unknown" – unresolved cell reference
                """
                if (sheet, row, col) in _xbrl_cell_map:
                    return _xbrl_cell_map[(sheet, row, col)], "xbrl"
                if (sheet, row, col) in _yellow_cell_set:
                    # Column-aware upward scan: prefer a label in the same
                    # column as the yellow cell.  Falls back to any-column
                    # row-level label if no same-column match is found.
                    tl: str | None = None
                    # 1) Same-column label at same row or up to 3 rows above
                    for look_back in range(0, 4):
                        tl = _vr_col_label.get((sheet, row - look_back, col))
                        if tl:
                            break
                    # 2) Fallback: any-column row-level label
                    if not tl:
                        tl = _vr_row_to_label.get((sheet, row))
                    if not tl:
                        for look_back in range(1, 4):
                            tl = _vr_row_to_label.get((sheet, row - look_back))
                            if tl:
                                break
                    if tl:
                        return tl, "yellow"
                    return f"yellow_{sheet}_{row}_{get_column_letter(col)}", "yellow"
                return None, "unknown"

            def _extract_comparison_values(
                formula: str,
                cell_col_letter: str,
                cell_row: int,
                current_sheet: str,
                primary_xbrl: str | None = None,
            ) -> dict[tuple[str, int, int], list[str]]:
                """Extract comparison values for cell references in the formula.

                For each cell reference, find patterns like
                ``CELL=VALUE``, ``CELL=TRUE``, ``CELL="text"`` and return a
                mapping from (sheet, row, col) → list of compared values.
                Only comparisons using ``=`` (equality) are captured;
                ``<>`` and ``""`` (empty-string checks) are skipped as those
                are self-reference "is field filled?" checks.

                Same-row references are included only when they resolve to a
                different identifier than *primary_xbrl* (i.e. they are
                condition triggers, not self-checks).
                """
                result: dict[tuple[str, int, int], list[str]] = {}

                def _is_self_ref(sh: str, rn: int, ci: int) -> bool:
                    """Return True if the cell is a self-reference to the
                    primary XBRL field being validated."""
                    if rn != cell_row:
                        return False
                    # Same row — resolve and check if it's the primary field
                    if primary_xbrl:
                        cid = _xbrl_cell_map.get((sh, rn, ci))
                        if cid == primary_xbrl:
                            return True  # actual self-ref
                        # Not the primary field — keep as potential condition
                        return False
                    # No primary_xbrl — fall back to skipping all same-row
                    return True

                # Collect all cell references with their positions in the formula
                # Cross-sheet: 'Sheet'!$COL$ROW
                for m in re.finditer(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)", formula):
                    sh, cl, rn = m.group(1), m.group(2), int(m.group(3))
                    ci = _col_to_idx(cl.replace("$", ""))
                    if _is_self_ref(sh, rn, ci):
                        continue
                    key = (sh, rn, ci)
                    # Look for =VALUE after the reference
                    after = formula[m.end() :]
                    eq_m = re.match(
                        r'\s*=\s*("([^"]*)"|TRUE|FALSE|[A-Za-z0-9_ ]+(?:\([^)]*\))?)',
                        after,
                    )
                    if eq_m:
                        val = (
                            eq_m.group(2)
                            if eq_m.group(2) is not None
                            else eq_m.group(1)
                        )
                        # Skip empty-string comparisons ("is field filled?" checks)
                        if val == "":
                            continue
                        result.setdefault(key, []).append(val)

                # Local: $COL$ROW or COL ROW (no sheet prefix).
                # Skip refs that are part of a cross-sheet reference
                # (preceded by '!' within 5 chars).
                for m in re.finditer(r"(?<!')(?<!\w)\$?([A-Z]{1,2})\$?(\d+)", formula):
                    # Check if this ref is preceded by '!' (cross-sheet)
                    _prefix = formula[max(0, m.start() - 2) : m.start()]
                    if "!" in _prefix:
                        continue
                    cl, rn = m.group(1), int(m.group(2))
                    ci = _col_to_idx(cl.replace("$", ""))
                    if _is_self_ref(current_sheet, rn, ci):
                        continue
                    key = (current_sheet, rn, ci)
                    after = formula[m.end() :]
                    eq_m = re.match(
                        r'\s*=\s*("([^"]*)"|TRUE|FALSE|[A-Za-z0-9_ ]+(?:\([^)]*\))?)',
                        after,
                    )
                    if eq_m:
                        val = (
                            eq_m.group(2)
                            if eq_m.group(2) is not None
                            else eq_m.group(1)
                        )
                        # Skip empty-string comparisons ("is field filled?" checks)
                        if val == "":
                            continue
                        result.setdefault(key, []).append(val)

                # Also detect direct named-range comparisons like
                # BasisForPreparation="Option A ..."
                for dn_name in _xbrl_defined_names:
                    for m in re.finditer(
                        re.escape(dn_name) + r'\s*=\s*("([^"]*)"|TRUE|FALSE)', formula
                    ):
                        val = m.group(2) if m.group(2) is not None else m.group(1)
                        if val == "":
                            continue
                        # Resolve the named range to its cell position
                        for _dn in wb_formulas.defined_names.values():
                            if _dn.name == dn_name:
                                for _dsh, _dco in _dn.destinations:
                                    try:
                                        _dcr = CellRange(_dco)
                                        _dk = (_dsh, _dcr.min_row, _dcr.min_col)
                                        result.setdefault(_dk, []).append(val)
                                    except Exception:
                                        pass
                                break

                return result

            for sheet_name in data_sheet_names:
                if sheet_name not in wb_formulas:
                    continue
                ws = wb_formulas[sheet_name]
                for row in ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row,
                    min_col=1,
                    max_col=ws.max_column,
                ):
                    for cell in row:
                        # Extract the formula text.  Regular formula cells
                        # store a plain str; array-formula cells wrap it in
                        # an ``ArrayFormula`` object whose ``.text``
                        # attribute holds the actual formula string.
                        _raw = cell.value
                        if isinstance(_raw, ArrayFormula):
                            _raw = _raw.text
                        if (
                            not _raw
                            or not isinstance(_raw, str)
                            or "template_label_ok" not in _raw
                        ):
                            continue

                        formula = _raw
                        has_missing = "template_label_missing_value" in formula

                        # Find the primary XBRL field on the same row.
                        # First try to the left (most common layout), then to
                        # the right for sheets where the validation column
                        # precedes the data column.
                        primary_xbrl: str | None = None
                        for c in range(1, cell.column):
                            key = (sheet_name, cell.row, c)
                            if key in _xbrl_cell_map:
                                primary_xbrl = _xbrl_cell_map[key]
                                break
                        if not primary_xbrl:
                            for c in range(cell.column + 1, ws.max_column + 1):
                                key = (sheet_name, cell.row, c)
                                if key in _xbrl_cell_map:
                                    primary_xbrl = _xbrl_cell_map[key]
                                    break
                        # Last resort: resolve cell references inside the
                        # formula to find the XBRL field being validated.
                        if not primary_xbrl:
                            _local_refs = re.findall(
                                r"(?<!')(?<!\w)\$?([A-Z]{1,2})\$?(\d+)",
                                formula,
                            )
                            for _cl, _rn in _local_refs:
                                _ci = _col_to_idx(_cl.replace("$", ""))
                                _xn = _xbrl_cell_map.get((sheet_name, int(_rn), _ci))
                                if _xn:
                                    primary_xbrl = _xn
                                    break

                        # Final fallback: derive an identifier from the
                        # template_label_* name in column C of the same row.
                        # This captures validation for meta / infrastructure
                        # fields (e.g. company name, dates, social metrics)
                        # that don't carry an XBRL defined name.
                        if not primary_xbrl:
                            _label_cell = ws.cell(row=cell.row, column=3)
                            if _label_cell.value and isinstance(_label_cell.value, str):
                                _lm = re.match(
                                    r"=?(template_label_\w+)",
                                    _label_cell.value,
                                )
                                if _lm:
                                    primary_xbrl = _lm.group(1)

                        if not primary_xbrl:
                            continue
                        # Only record the first validation cell per XBRL field
                        if primary_xbrl in validation_rules:
                            continue

                        # --- Detect conditions --------------------------------
                        # Extract ALL cell references from the formula that
                        # point to a *different row* from the validation cell.
                        # Same-row references are "is field empty?" self-checks.
                        # Resolve each to an XBRL name, yellow-cell label, or
                        # skip if unresolvable.
                        #
                        # Additionally resolve XBRL defined names that appear
                        # verbatim in the formula (e.g. BasisForPreparation).
                        condition_ids: dict[str, str] = {}
                        # identifier → category ("xbrl" | "yellow")

                        # (a) Direct named-range references in formula
                        for dn_name in _xbrl_defined_names:
                            if dn_name != primary_xbrl and dn_name in formula:
                                condition_ids[dn_name] = "xbrl"

                        # (b) Cross-sheet cell references
                        cross = re.findall(r"'([^']+)'!\$?([A-Z]+)\$?(\d+)", formula)
                        for cs, cl, rn in cross:
                            ci = _col_to_idx(cl.replace("$", ""))
                            ri = int(rn)
                            if ri == cell.row and cs == sheet_name:
                                # Same row — only skip if it resolves to the
                                # primary XBRL field (true self-reference).
                                _sr_id = _xbrl_cell_map.get((cs, ri, ci))
                                if _sr_id == primary_xbrl:
                                    continue
                            cid, cat = _resolve_cell_id(cs, ri, ci)
                            if cid and cid != primary_xbrl:
                                condition_ids.setdefault(cid, cat)

                        # (c) Local cell references (possibly same row).
                        #     Skip refs that are part of a cross-sheet reference
                        #     (preceded by '!' within 2 chars).
                        #     Same-row refs are included only if they resolve to
                        #     a different identifier than primary_xbrl (i.e. they
                        #     are condition triggers, not self-checks).
                        for _lr_m in re.finditer(
                            r"(?<!')(?<!\w)\$?([A-Z]{1,2})\$?(\d+)", formula
                        ):
                            _pref = formula[max(0, _lr_m.start() - 2) : _lr_m.start()]
                            if "!" in _pref:
                                continue
                            cl, rn = _lr_m.group(1), _lr_m.group(2)
                            ci = _col_to_idx(cl.replace("$", ""))
                            ri = int(rn)
                            if ri == cell.row:
                                # Same row — only skip true self-references
                                _sr_id = _xbrl_cell_map.get((sheet_name, ri, ci))
                                if _sr_id == primary_xbrl:
                                    continue
                            cid, cat = _resolve_cell_id(sheet_name, ri, ci)
                            if cid and cid != primary_xbrl:
                                condition_ids.setdefault(cid, cat)

                        # --- Extract comparison values for conditions ---------
                        comparison_values = _extract_comparison_values(
                            formula,
                            get_column_letter(cell.column),
                            cell.row,
                            sheet_name,
                            primary_xbrl=primary_xbrl,
                        )

                        # Build condition_criteria: for each condition variable,
                        # collect the comparison values from the formula.
                        # Deduplicate and join with "|" (pipe) for the
                        # condition-eval truth_map format.
                        condition_criteria: dict[str, str] = {}
                        for cid in condition_ids:
                            # Find comparison values for this condition ID
                            cid_values: list[str] = []
                            # Match by resolving all cell positions for this ID
                            for (cs, cr, cc), vals in comparison_values.items():
                                resolved_id, _ = _resolve_cell_id(cs, cr, cc)
                                # Also check XBRL cell map directly
                                if resolved_id is None:
                                    resolved_id = _xbrl_cell_map.get((cs, cr, cc))
                                if resolved_id == cid:
                                    cid_values.extend(vals)
                            # Deduplicate while preserving order
                            seen: set[str] = set()
                            unique_vals: list[str] = []
                            for v in cid_values:
                                if v not in seen:
                                    seen.add(v)
                                    unique_vals.append(v)
                            if unique_vals:
                                condition_criteria[cid] = "|".join(unique_vals)

                        # Build the condition struct_expr in {var} format.
                        # All condition variables are AND-ed together.
                        condition_expr: str | None = None
                        if condition_ids:
                            parts = [f"{{{cid}}}" for cid in sorted(condition_ids)]
                            if len(parts) == 1:
                                condition_expr = parts[0]
                            else:
                                condition_expr = "&".join(parts)

                        # Pick the most meaningful condition field for the
                        # legacy conditionField key.  Prefer a field that is
                        # NOT BasisForPreparation when other specific
                        # conditions exist alongside it.
                        condition_xbrl: str | None = None
                        if condition_ids:
                            non_generic = {
                                k for k in condition_ids if k != "BasisForPreparation"
                            }
                            if non_generic:
                                condition_xbrl = sorted(non_generic)[0]
                            else:
                                condition_xbrl = sorted(condition_ids)[0]

                        rule: dict = {
                            "required": has_missing,
                            "validationCell": f"{sheet_name}!{cell.coordinate}",
                        }
                        if condition_xbrl:
                            rule["conditionField"] = condition_xbrl
                        if len(condition_ids) > 1:
                            rule["allConditionFields"] = sorted(condition_ids)
                        if condition_expr:
                            rule["condition"] = condition_expr
                        if condition_criteria:
                            rule["conditionCriteria"] = condition_criteria
                        if not has_missing:
                            rule["status"] = "informational"
                        elif condition_ids:
                            rule["status"] = "conditional"
                        else:
                            rule["status"] = "required"

                        validation_rules[primary_xbrl] = rule

            # --- Template warnings -------------------------------------------
            # Some cells adjacent to validation cells (typically one column to
            # the right) contain references to ``template_label_*_warning``
            # names.  These are informational hints displayed to the user (e.g.
            # "please select a NACE code rather than a category").
            #
            # We scan for simple ``=template_label_*_warning`` references and
            # associate each with the XBRL field (or template_label fallback)
            # on the same row, mirroring the logic used for validation cells.

            _WARNING_RE = re.compile(r"^=?(template_label_\w*warning\w*)$")

            for sheet_name in data_sheet_names:
                if sheet_name not in wb_formulas:
                    continue
                ws = wb_formulas[sheet_name]
                for row_cells in ws.iter_rows(
                    min_row=1,
                    max_row=ws.max_row,
                    min_col=1,
                    max_col=ws.max_column,
                ):
                    for cell in row_cells:
                        if not cell.value or not isinstance(cell.value, str):
                            continue
                        wm = _WARNING_RE.match(cell.value.strip())
                        if not wm:
                            continue
                        warning_name = wm.group(1)

                        # Find the field on the same row (left, right, formula fallback)
                        field_id: str | None = None
                        for c in range(1, cell.column):
                            key = (sheet_name, cell.row, c)
                            if key in _xbrl_cell_map:
                                field_id = _xbrl_cell_map[key]
                                break
                        if not field_id:
                            for c in range(cell.column + 1, ws.max_column + 1):
                                key = (sheet_name, cell.row, c)
                                if key in _xbrl_cell_map:
                                    field_id = _xbrl_cell_map[key]
                                    break
                        # Fallback: template_label from col C
                        if not field_id:
                            _lc = ws.cell(row=cell.row, column=3)
                            if _lc.value and isinstance(_lc.value, str):
                                _lm2 = re.match(r"=?(template_label_\w+)", _lc.value)
                                if _lm2 and "warning" not in _lm2.group(1):
                                    field_id = _lm2.group(1)
                        if not field_id:
                            continue

                        # Build the warning key from the name, stripping
                        # the ``template_label_`` prefix.
                        warning_key = warning_name[len("template_label_") :]
                        colL = get_column_letter(cell.column)
                        warning_entry = {
                            "warningName": warning_name,
                            "warningKey": warning_key,
                            "warningCell": f"{sheet_name}!{colL}{cell.row}",
                        }

                        # Attach to the validation rule for this field, or
                        # create a minimal validation entry with just the
                        # warning if no rule exists yet.
                        if field_id in validation_rules:
                            existing = validation_rules[field_id]
                            if "warnings" not in existing:
                                existing["warnings"] = []
                            # Avoid duplicates
                            if warning_entry not in existing["warnings"]:
                                existing["warnings"].append(warning_entry)
                        else:
                            validation_rules[field_id] = {
                                "required": False,
                                "status": "informational",
                                "warnings": [warning_entry],
                            }

            # --- Section header labels from formulas -------------------------
            # Each section ``template_*`` defined name points to a header
            # cell whose formula concatenates the section label with date
            # range and applicability strings, e.g.:
            #   =template_label_b2_cooperative_specific_disclosures & " "
            #     & template_label_from & …
            # We extract the *first* ``template_label_*`` identifier from
            # that formula and look it up in the already-resolved ``labels``
            # dict to get the clean, human-readable header for each section.
            _section_label_re = re.compile(r"template_label_(\w+)")
            _non_section = {
                "template_currency",
                "template_selected_display_language",
                "template_overall_validation_status",
                "template_starting_date_display",
                "template_translations",
            }
            for _dn in wb_formulas.defined_names.values():
                if not _dn.name or not _dn.name.startswith("template_"):
                    continue
                if _dn.name.startswith("template_label_") or _dn.name.startswith(
                    "template_reporting_"
                ):
                    continue
                if _dn.name in _non_section:
                    continue
                # Read the raw formula from the first destination cell
                for _sh, _co in _dn.destinations:
                    if _sh not in wb_formulas:
                        continue
                    try:
                        _cr = CellRange(_co)
                    except Exception:
                        continue
                    _cell = wb_formulas[_sh].cell(row=_cr.min_row, column=_cr.min_col)
                    _formula = str(_cell.value or "")
                    _m = _section_label_re.search(_formula)
                    if _m:
                        _lkey = _m.group(1)  # key into ``labels``
                        _header = labels.get(_lkey)
                        if _header and isinstance(_header, str):
                            # Store keyed by the part after ``template_``
                            _section_key = _dn.name[len("template_") :]
                            section_labels[_section_key] = _header
                    break

            wb_formulas.close()

        # --- Excel-native sections -----------------------------------------
        # Build the definitive section list directly from ``template_*``
        # defined names that point to header cells in the four data sheets.
        # Each header cell's formula references an applicability label
        # (template_label_always_to_be_reported, template_label_if_applicable,
        # etc.) from which we derive the section's applicability status.
        # XBRL fields are assigned to sections by their cell row falling
        # within the section's row range on the same sheet (+ column range
        # for the side-by-side GHG / GHG-targets case).

        _DATA_SHEET_NAMES = [
            "General Information",
            "Environmental Disclosures",
            "Social Disclosures",
            "Governance Disclosures",
        ]
        _SHEET_ORDER = {s: i for i, s in enumerate(_DATA_SHEET_NAMES)}

        # Names that live in data sheets but are NOT section headers
        _NON_SECTION_NAMES = {
            "template_currency",
            "template_selected_display_language",
            "template_overall_validation_status",
            "template_starting_date_display",
            "template_translations",
        }

        raw_headers: list[dict] = []
        for dn in self._workbook.defined_names.values():
            if not dn.name or not dn.name.startswith("template_"):
                continue
            if dn.name.startswith("template_label_") or dn.name.startswith(
                "template_reporting_"
            ):
                continue
            if dn.name in _NON_SECTION_NAMES:
                continue
            for sheet, coord in dn.destinations:
                if sheet not in _SHEET_ORDER:
                    continue
                try:
                    cr = CellRange(coord)
                except Exception:
                    continue
                row, col = cr.min_row, cr.min_col
                ws = self._workbook[sheet]
                cell = ws.cell(row=row, column=col)
                is_bold = cell.font and cell.font.bold
                if not is_bold:
                    continue  # not a visual section header

                # Determine applicability from the resolved header text.
                # The workbook is loaded with data_only=True so cell.value
                # contains the evaluated display string, not the formula.
                appl = "always"  # default for headers without an explicit flag
                header_text = str(cell.value) if cell.value else ""
                if "[Always to be reported + If applicable]" in header_text:
                    appl = "always+conditional"
                elif "[If applicable linked with" in header_text:
                    appl = "conditional-linked"
                elif "[If applicable]" in header_text:
                    appl = "conditional"
                elif "[May (optional)]" in header_text:
                    appl = "optional"
                elif "[Always to be reported]" in header_text:
                    appl = "always"

                raw_headers.append(
                    {
                        "name": dn.name,
                        "sheet": sheet,
                        "row": row,
                        "col": col,
                        "applicability": appl,
                    }
                )
                break  # only first destination

        # Sort by sheet order, then row, then column
        raw_headers.sort(
            key=lambda h: (_SHEET_ORDER.get(h["sheet"], 99), h["row"], h["col"])
        )

        # Build (sheet, row, col) → set[XBRL-name] map for field assignment.
        # Multiple defined names can overlap on the same cells (e.g. a Table
        # name covering the same range as individual field names), so we
        # store a *set* of names per cell position.
        _xbrl_cell_positions: dict[tuple[str, int, int], set[str]] = {}
        for dn in self._workbook.defined_names.values():
            if not dn.name or dn.name.startswith(("template_", "enum_")):
                continue
            for _sh, _co in dn.destinations:
                if _sh not in _SHEET_ORDER:
                    continue
                try:
                    _cr = CellRange(_co)
                except Exception:
                    continue
                for _r in range(_cr.min_row, _cr.max_row + 1):
                    for _c in range(_cr.min_col, _cr.max_col + 1):
                        _xbrl_cell_positions.setdefault((_sh, _r, _c), set()).add(
                            dn.name
                        )

        # --- Input-cell helper fields ---------------------------------------------
        # Cells with an input-cell fill colour (yellow FFFFFF99 or light-gray
        # theme:0/tint:-0.15) are helper/condition fields that are not part
        # of the XBRL report but facilitate completion of the disclosures.
        # They include boolean trigger questions ("Has the undertaking…?"),
        # helper numeric inputs (e.g. board-member counts that feed calculated
        # XBRL ratios), and dropdown selections.
        #
        # We scan for input cells that carry a user-editable value (boolean,
        # number, or dropdown), derive a stable identifier from the
        # ``template_label_*`` formula on the same row, and collect their
        # metadata so they can be included in section field lists.
        #
        # We use the formula workbook (wb_formulas) to read the raw formulas
        # so we can identify label references; the value workbook gives the
        # evaluated cell values and fill colours.

        # yellow_field_id → descriptor dict
        _yellow_fields: dict[str, dict] = {}
        # (sheet, row, col) → yellow_field_id  (for value cells only)
        _yellow_cell_positions: dict[tuple[str, int, int], str] = {}

        # Positions already covered by XBRL defined names – skip those
        _xbrl_occupied: set[tuple[str, int, int]] = set(_xbrl_cell_positions.keys())

        for sheet_name in _DATA_SHEET_NAMES:
            if sheet_name not in self._workbook:
                continue
            ws_val = self._workbook[sheet_name]
            ws_form = (
                wb_formulas[sheet_name]
                if (wb_formulas is not None and sheet_name in wb_formulas)
                else None
            )

            # First pass: collect all input cells on this sheet
            # (yellow fill FFFFFF99 or light-gray theme:0/tint:-0.15)
            yellow_on_sheet: dict[int, dict[int, object]] = {}  # row → {col → value}
            # Also collect computed cells (theme:8/tint≈0.80) on the same sheet
            computed_on_sheet: dict[int, dict[int, object]] = {}  # row → {col → value}
            for row_cells in ws_val.iter_rows(
                min_row=1,
                max_row=ws_val.max_row,
                min_col=1,
                max_col=ws_val.max_column,
            ):
                for cell in row_cells:
                    if _is_input_cell_fill(cell):
                        yellow_on_sheet.setdefault(cell.row, {})[
                            cell.column
                        ] = cell.value
                    elif _is_computed_cell_fill(cell):
                        computed_on_sheet.setdefault(cell.row, {})[
                            cell.column
                        ] = cell.value

            # Second pass: for each row with yellow cells, identify the label
            # cell (template_label_* formula in col C or nearby) and the value
            # cell(s).
            _LABEL_RE = re.compile(r"^=?(template_label_\w+)")

            for row_num, col_vals in yellow_on_sheet.items():
                # Find the label key from the template_label_* formula on this
                # row.  The label is usually in column C (3) but can be in any
                # yellow cell that contains a template_label_ reference.
                label_key: str | None = None
                label_text: str | None = None
                template_label_full: str | None = (
                    None  # full template_label_* defined name
                )
                for c in sorted(col_vals.keys()):
                    if ws_form is not None:
                        form_val = ws_form.cell(row=row_num, column=c).value
                        if isinstance(form_val, str):
                            m = _LABEL_RE.match(form_val.strip())
                            if m:
                                full_name = m.group(
                                    1
                                )  # e.g. template_label_has_the_undertaking_…
                                label_key = full_name[len("template_label_") :]
                                template_label_full = full_name
                                # Resolved display text from the value workbook
                                label_text = (
                                    str(col_vals[c])
                                    if col_vals[c] is not None
                                    else None
                                )
                                break

                if not label_key and ws_form is not None:
                    # No label found on the same row.  Look upward in the same
                    # column (within 3 rows) for an input-fill label cell —
                    # this covers multi-row question layouts (e.g. GI row 554
                    # label / row 556 answer grid, or GI row 447 label / 449
                    # answer).
                    # Use the smallest (leftmost) answer column as the anchor.
                    anchor_col = min(col_vals.keys())
                    for look_back in range(1, 4):
                        candidate_row = row_num - look_back
                        if candidate_row < 1:
                            break
                        cand_cell_form = ws_form.cell(
                            row=candidate_row, column=anchor_col
                        )
                        cand_cell_val = ws_val.cell(
                            row=candidate_row, column=anchor_col
                        )
                        if _is_input_cell_fill(cand_cell_val) and isinstance(
                            cand_cell_form.value, str
                        ):
                            m = _LABEL_RE.match(cand_cell_form.value.strip())
                            if m:
                                full_name = m.group(1)
                                label_key = full_name[len("template_label_") :]
                                template_label_full = full_name
                                # Resolved display text from the value workbook
                                lv = ws_val.cell(
                                    row=candidate_row, column=anchor_col
                                ).value
                                label_text = str(lv) if lv is not None else None
                                break

                if not label_key:
                    # No identifiable label on this row or above — skip
                    continue

                # Find value cells: yellow cells on this row that are NOT
                # the label cell and NOT already covered by XBRL names.
                # Also skip cells whose formulas are complex calculations
                # (they reference other cells with arithmetic operators).
                _CALC_RE = re.compile(r"[+\-*/]")

                for c, val in sorted(col_vals.items()):
                    # Skip if this is the label cell
                    if ws_form is not None:
                        form_val = ws_form.cell(row=row_num, column=c).value
                        if isinstance(form_val, str) and _LABEL_RE.match(
                            form_val.strip()
                        ):
                            continue
                        # Skip computed formula cells (contain arithmetic)
                        if (
                            isinstance(form_val, str)
                            and form_val.startswith("=")
                            and _CALC_RE.search(form_val)
                        ):
                            continue

                    # Skip cells already covered by XBRL defined names
                    if (sheet_name, row_num, c) in _xbrl_occupied:
                        continue

                    # Skip cells with only label text and no input value
                    # (e.g. Scope 3 category names in col C rows 30-44)
                    if val is None:
                        continue

                    # Determine the input type from the value and data validation
                    input_type = "text"
                    data_type: str | None = None
                    enum_ref: str | None = None

                    if isinstance(val, bool):
                        input_type = "boolean"
                        data_type = "booleanItemType"
                    elif isinstance(val, (int, float)):
                        input_type = "number"
                        data_type = "decimalItemType"
                    elif isinstance(val, (str, CellRichText)) and (
                        str(val) == "-" or str(val) == "\u2013"
                    ):
                        continue  # dash = empty computed cell, skip
                    elif isinstance(val, (str, CellRichText)):
                        # Preliminary inputType; confirmed as "select" below if a
                        # data-validation dropdown is found.  Pure static text cells
                        # (e.g. privacy-policy notices) that have no dropdown are
                        # excluded — they are display-only, not user-editable inputs.
                        input_type = "text"

                    # Check data validation for dropdown
                    if ws_val.data_validations:
                        for dv in ws_val.data_validations.dataValidation:
                            if dv.formula1 and dv.formula1.startswith("enum_"):
                                for sqref_str in str(dv.sqref).split():
                                    try:
                                        dv_cr = CellRange(sqref_str)
                                        if (
                                            dv_cr.min_row <= row_num <= dv_cr.max_row
                                            and dv_cr.min_col <= c <= dv_cr.max_col
                                        ):
                                            enum_ref = dv.formula1.strip()
                                            input_type = "select"
                                            data_type = "enumerationItemType"
                                    except Exception:
                                        pass

                    # Skip static text cells: plain string (or rich text) with
                    # no dropdown (e.g. privacy-policy notice in GI!H448).
                    if input_type == "text" and isinstance(val, (str, CellRichText)):
                        continue

                    col_letter = get_column_letter(c)
                    field_id = f"yellow_{label_key}_{col_letter}{row_num}"

                    # i18n labels from translations dict
                    field_labels: dict[str, str] = {}
                    if label_key in translations:
                        field_labels = dict(translations[label_key])
                    elif label_text:
                        field_labels["en"] = label_text

                    descriptor: dict = {
                        "fieldId": field_id,
                        "source": "yellowCell",
                        "labelKey": label_key,
                        "templateLabelKey": template_label_full,
                        "label": label_text or label_key,
                        "labels": field_labels,
                        "inputType": input_type,
                        "dataType": data_type,
                        "value": fact_value_to_json(val),
                        "cellRef": f"{sheet_name}!{col_letter}{row_num}",
                        "sheet": sheet_name,
                        "row": row_num,
                        "col": c,
                        "isRequired": False,
                        "isReportable": False,
                    }
                    if enum_ref:
                        descriptor["options"] = enum_ref

                    _yellow_fields[field_id] = descriptor
                    _yellow_cell_positions[(sheet_name, row_num, c)] = field_id

                # --- Computed cells on the same row as a yellow label --------
                # Some rows have a yellow label cell (C) with a computed value
                # cell (D) — e.g. "Total hours worked" in b9 row 142.
                # These need a field entry marked isComputed=True.
                # Only pick up computed cells whose column is adjacent to the
                # yellow cell column range on this row — this avoids false
                # positives from side-by-side sections (e.g. b3/c3 share rows).
                if label_key and row_num in computed_on_sheet:
                    _min_yc = min(col_vals.keys())
                    _max_yc = max(col_vals.keys())
                    for c, val in sorted(computed_on_sheet[row_num].items()):
                        # Must be within ±1 of the yellow cell column range
                        if c < _min_yc - 1 or c > _max_yc + 1:
                            continue
                        # Skip label cells
                        if ws_form is not None:
                            form_val = ws_form.cell(row=row_num, column=c).value
                            if isinstance(form_val, str) and _LABEL_RE.match(
                                form_val.strip()
                            ):
                                continue
                        # Skip cells already covered by XBRL defined names
                        if (sheet_name, row_num, c) in _xbrl_occupied:
                            continue
                        # Skip dashes (empty computed cells)
                        if val is None:
                            continue
                        if isinstance(val, (str, CellRichText)) and (
                            str(val) == "-" or str(val) == "\u2013"
                        ):
                            continue

                        # Determine input type from value
                        comp_input_type = "number"
                        comp_data_type = "decimalItemType"
                        if isinstance(val, bool):
                            comp_input_type = "boolean"
                            comp_data_type = "booleanItemType"
                        elif isinstance(val, (str, CellRichText)):
                            comp_input_type = "text"
                            comp_data_type = None

                        col_letter = get_column_letter(c)
                        comp_field_id = f"yellow_{label_key}_{col_letter}{row_num}"

                        # Skip if we already created a yellow field for this cell
                        if comp_field_id in _yellow_fields:
                            continue

                        # i18n labels
                        comp_labels: dict[str, str] = {}
                        if label_key in translations:
                            comp_labels = dict(translations[label_key])
                        elif label_text:
                            comp_labels["en"] = label_text

                        comp_descriptor: dict = {
                            "fieldId": comp_field_id,
                            "source": "yellowCell",
                            "labelKey": label_key,
                            "templateLabelKey": template_label_full,
                            "label": label_text or label_key,
                            "labels": comp_labels,
                            "inputType": comp_input_type,
                            "dataType": comp_data_type,
                            "value": fact_value_to_json(val),
                            "cellRef": f"{sheet_name}!{col_letter}{row_num}",
                            "sheet": sheet_name,
                            "row": row_num,
                            "col": c,
                            "isRequired": False,
                            "isReportable": False,
                            "isComputed": True,
                        }

                        _yellow_fields[comp_field_id] = comp_descriptor
                        _yellow_cell_positions[(sheet_name, row_num, c)] = comp_field_id

        # --- Row → template_label mapping ------------------------------------
        # For every data-sheet row that contains a ``=template_label_*``
        # formula we record the *leftmost* label key.  This lets us attach
        # the spreadsheet's question phrasing (and its i18n translations)
        # to both XBRL and yellow-cell fields.
        _LABEL_RE_TL = re.compile(r"^=?(template_label_\w+)")

        # Comprehensive cell-level map:
        #   (sheet, row, col) → (template_label_name, is_bold, merge_col_span)
        # This captures EVERY template_label formula cell for matrix detection.
        # The ``is_bold`` flag indicates whether the cell uses bold font,
        # which distinguishes column headers (bold) from row labels (non-bold).
        # ``merge_col_span`` is the number of columns the cell's merged range
        # covers (1 = not merged).  Wide-merged bold cells are section-wide
        # headers rather than table column headers.
        _label_cell_map: dict[tuple[str, int, int], tuple[str, bool, int]] = {}
        _row_to_template_label: dict[tuple[str, int], str] = {}
        #   (sheet_name, row) → label_key (without the "template_label_" prefix)

        # ``additional_rows_warning`` cells mark where the repeatable
        # table data area ends.  Fields below this row are summary/total
        # rows that should not inherit table column labels.
        # Keyed by (sheet_name, row) with the column index as value.
        _additional_rows_warnings: dict[tuple[str, int], int] = {}

        if wb_formulas is not None:
            for sheet_name in _DATA_SHEET_NAMES:
                if sheet_name not in wb_formulas:
                    continue
                ws_f = wb_formulas[sheet_name]
                # The data-only workbook preserves cell formatting (font, fill)
                # which we need to detect bold column headers.
                ws_val = (
                    self._workbook[sheet_name] if sheet_name in self._workbook else None
                )
                # Build a lookup for merged cell column spans so we can
                # distinguish narrow column headers from wide section-wide
                # instruction rows.
                _merge_span: dict[tuple[int, int], int] = {}  # (row, col) → col_span
                if ws_val is not None:
                    for mr in ws_val.merged_cells.ranges:
                        span = mr.max_col - mr.min_col + 1
                        _merge_span[(mr.min_row, mr.min_col)] = span

                for row_cells in ws_f.iter_rows(
                    min_row=1,
                    max_row=ws_f.max_row,
                    min_col=1,
                    max_col=ws_f.max_column,
                ):
                    for cell in row_cells:
                        if not isinstance(cell.value, str):
                            continue
                        m = _LABEL_RE_TL.match(cell.value.strip())
                        if not m:
                            continue
                        full_name = m.group(1)  # e.g. "template_label_foo"
                        lk = full_name[len("template_label_") :]
                        # Store every label cell for matrix detection,
                        # together with the bold status read from the
                        # value workbook which preserves formatting.
                        if "warning" not in full_name:
                            is_bold = False
                            if ws_val is not None:
                                val_cell = ws_val.cell(row=cell.row, column=cell.column)
                                is_bold = bool(val_cell.font and val_cell.font.bold)
                            col_span = _merge_span.get((cell.row, cell.column), 1)
                            _label_cell_map[(sheet_name, cell.row, cell.column)] = (
                                full_name,
                                is_bold,
                                col_span,
                            )
                        elif full_name == "template_label_additional_rows_warning":
                            _additional_rows_warnings[(sheet_name, cell.row)] = (
                                cell.column
                            )
                            # Warning cells must NOT populate
                            # _row_to_template_label — otherwise
                            # range-based XBRL names spanning the
                            # warning row would inherit
                            # "additional_rows_warning" as their
                            # labelKey.
                            continue
                        key = (sheet_name, cell.row)
                        if key not in _row_to_template_label:
                            # First (leftmost) label wins — later labels
                            # on the same row are warnings / guidance.
                            _row_to_template_label[key] = lk

        # Map XBRL defined-name → template_label_key using their cell rows.
        _xbrl_name_to_template_label: dict[str, str] = {}
        for (sh, r, _c), names in _xbrl_cell_positions.items():
            tl_key = _row_to_template_label.get((sh, r))
            if tl_key:
                for name in names:
                    if name not in _xbrl_name_to_template_label:
                        _xbrl_name_to_template_label[name] = tl_key

        # Compute row/col boundaries for each section and assign XBRL fields
        excel_sections: list[dict] = []
        for idx, hdr in enumerate(raw_headers):
            # Determine the row boundary: up to the next header on same sheet
            # (respecting column for the side-by-side GHG case on row 16)
            next_row = None
            sheet_max_row = self._workbook[hdr["sheet"]].max_row
            for later in raw_headers[idx + 1 :]:
                if later["sheet"] != hdr["sheet"]:
                    break
                # If the next header is on the same row but different column
                # (side-by-side sections), don't treat it as a row boundary
                if later["row"] == hdr["row"] and later["col"] != hdr["col"]:
                    continue
                next_row = later["row"]
                break
            end_row = (next_row - 1) if next_row else sheet_max_row

            # For side-by-side sections on the same row, determine column range
            # Check if there's another header on the same row
            same_row_headers = [
                h
                for h in raw_headers
                if h["sheet"] == hdr["sheet"] and h["row"] == hdr["row"]
            ]
            col_min = hdr["col"]
            col_max = self._workbook[hdr["sheet"]].max_column
            if len(same_row_headers) > 1:
                # Sort by column
                same_row_headers.sort(key=lambda h: h["col"])
                my_idx = next(
                    i
                    for i, h in enumerate(same_row_headers)
                    if h["name"] == hdr["name"]
                )
                if my_idx + 1 < len(same_row_headers):
                    col_max = same_row_headers[my_idx + 1]["col"] - 1

            # Find XBRL fields within this section's cell range,
            # ordered by row then column (top-to-bottom, left-to-right)
            # to match the visual order in the Excel sheet.
            field_positions: list[tuple[int, int, str]] = []
            seen: set[str] = set()
            for (_sh, _r, _c), xbrl_names_set in _xbrl_cell_positions.items():
                if _sh != hdr["sheet"]:
                    continue
                if not (hdr["row"] < _r <= end_row):  # fields below header row
                    continue
                if not (col_min <= _c <= col_max):
                    continue
                for xbrl_name in xbrl_names_set:
                    if xbrl_name not in seen:
                        seen.add(xbrl_name)
                        field_positions.append((_r, _c, xbrl_name))

            # Also include yellow-cell helper fields in this section's range
            yellow_fields_in_section: dict[str, dict] = {}
            for (_sh, _r, _c), yf_id in _yellow_cell_positions.items():
                if _sh != hdr["sheet"]:
                    continue
                if not (hdr["row"] < _r <= end_row):
                    continue
                if not (col_min <= _c <= col_max):
                    continue
                if yf_id not in seen:
                    seen.add(yf_id)
                    field_positions.append((_r, _c, yf_id))
                    yellow_fields_in_section[yf_id] = _yellow_fields[yf_id]

            field_positions.sort(key=lambda t: (t[0], t[1]))
            xbrl_fields: list[str] = [name for _, _, name in field_positions]

            # --- Collect unit-selection fields --------------------------------
            # Defined names ending in ``_unit`` are dropdown fields that let
            # the preparer choose the unit of measurement (e.g. kg vs tonne).
            # They are NOT taxonomy concepts and NOT yellow cells, so they
            # need to be handled as a third field category.
            _UNIT_RE = re.compile(r"_unit$")
            unit_fields_in_section: dict[str, dict] = {}
            for fn in xbrl_fields:
                if not _UNIT_RE.search(fn):
                    continue
                if fn in yellow_fields_in_section:
                    continue  # already handled as yellow cell
                pos = None
                for _r, _c, _name in field_positions:
                    if _name == fn:
                        pos = (_r, _c)
                        break
                if pos is None:
                    continue
                u_row, u_col = pos
                # Read cell value from the workbook
                ws_val = self._workbook[hdr["sheet"]]
                cell_val = ws_val.cell(row=u_row, column=u_col).value
                if cell_val is not None:
                    cell_val = fact_value_to_json(cell_val)
                # Find the template label for this field
                tl_key = _xbrl_name_to_template_label.get(fn)
                tl_labels: dict[str, str] = {}
                if tl_key and translations and tl_key in translations:
                    tl_labels = dict(translations[tl_key])
                # Build enum options from xbrl_to_enum
                options: list[str] = []
                if fn in xbrl_to_enum:
                    for enum_ref in xbrl_to_enum[fn]:
                        options.extend(enum_lists.get(enum_ref, []))
                # Build validation info
                vr = validation_rules.get(fn)
                # Cell reference for traceability
                cell_ref = f"{hdr['sheet']}!{get_column_letter(u_col)}{u_row}"
                unit_fields_in_section[fn] = {
                    "fieldId": fn,
                    "value": cell_val,
                    "labelKey": tl_key,
                    "templateLabelKey": (
                        f"template_label_{tl_key}" if tl_key else None
                    ),
                    "labels": tl_labels,
                    "options": options if options else None,
                    "validation": vr,
                    "cellRef": cell_ref,
                }

            # --- Build field position map: field_name → (row, col) -----------
            field_position_map: dict[str, tuple[int, int]] = {}
            for _r, _c, _name in field_positions:
                if _name not in field_position_map:
                    field_position_map[_name] = (_r, _c)

            # --- Detect matrix structure: row labels & column labels ---------
            # Classification uses **bold font** as the discriminator:
            #   • Bold ``=template_label`` cells → column / table headers
            #   • Non-bold ``=template_label`` cells → row labels / question labels
            # Section-title cells (on the exact header row, same column as the
            # defined-name anchor, usually also bold) are excluded because they
            # are section titles, not data-column headers.
            #
            # **Exception — computed-row labels:** Bold labels on rows that
            # contain a computed cell (theme tint ≈ 0.80) are treated as
            # row labels, not column headers.  These are auto-calculated
            # summary fields (e.g. accident rate, turnover rate, gender
            # ratio) whose label the template bolds for emphasis, but the
            # cell is not a table column header.

            # Collect all template_label cells within this section's bounds,
            # together with their bold status and merge column span.
            section_label_cells: list[tuple[int, int, str, bool, int]] = []
            # Identify rows with computed cells within this section
            _computed_rows_in_section: set[int] = set()
            for (_sh, _r, _c), (lbl_name, is_bold, col_span) in _label_cell_map.items():
                if _sh != hdr["sheet"]:
                    continue
                if not (hdr["row"] <= _r <= end_row):
                    continue
                if not (col_min <= _c <= col_max):
                    continue
                section_label_cells.append((_r, _c, lbl_name, is_bold, col_span))
            for _cr_sh, _cr_r, _cr_c in _computed_cell_set:
                if _cr_sh == hdr["sheet"] and hdr["row"] <= _cr_r <= end_row:
                    _computed_rows_in_section.add(_cr_r)
            section_label_cells.sort(key=lambda t: (t[0], t[1]))

            # Identify the section-title cell(s) to exclude from label
            # classification.  Section titles sit on the header row and their
            # name starts with ``template_label_`` followed by the section id
            # (minus the ``template_`` prefix).  In practice the title cell is
            # on hdr["row"] at hdr["col"].
            title_positions: set[tuple[int, int]] = set()
            title_positions.add((hdr["row"], hdr["col"]))

            # Row labels: NON-bold template_label cells (excluding titles),
            # plus bold labels on computed rows (auto-calculated fields).
            section_row_labels: dict[int, str] = {}  # row → template_label_name
            # Map each row label to its column so we can later identify
            # index columns (columns that contain row labels).
            _row_label_col_map: dict[int, int] = {}  # row → col
            for _r, _c, lbl_name, is_bold, _cs in section_label_cells:
                if (_r, _c) in title_positions:
                    continue
                # Bold labels on computed rows are treated as row labels
                effective_bold = is_bold and _r not in _computed_rows_in_section
                if not effective_bold:
                    if _r not in section_row_labels:
                        section_row_labels[_r] = lbl_name
                        _row_label_col_map[_r] = _c

            # Column labels: BOLD template_label cells (excluding titles)
            #
            # Wide-merged bold cells (spanning ≥ _WIDE_MERGE_THRESHOLD
            # columns) that are the *only* bold template_label on their row
            # are section-wide header rows (e.g. unit-selection instruction
            # rows), NOT table column headers.  They are stored separately
            # in ``section_wide_labels`` for unit-field label resolution.
            # Single-bold cells that are NOT wide-merged (e.g. a lone column
            # header like "Monetary amount in EUR") remain column labels.
            _WIDE_MERGE_THRESHOLD = 4  # columns; ≥ this → section-wide

            from collections import Counter as _Counter

            _bold_per_row: _Counter[int] = _Counter()
            for _r, _c, _lbl, _ib, _cs in section_label_cells:
                if (_r, _c) in title_positions:
                    continue
                # Bold on computed rows doesn't count as a column header
                if _ib and _r not in _computed_rows_in_section:
                    _bold_per_row[_r] += 1

            section_col_labels: dict[int, str] = {}  # col → template_label_name
            col_label_row: int | None = None  # row of the first column header
            # Section-wide labels: wide-merged bold cells on rows with only
            # one bold cell.  Keyed by row number.
            section_wide_labels: dict[int, str] = {}
            for _r, _c, lbl_name, is_bold, _cs in section_label_cells:
                if (_r, _c) in title_positions:
                    continue
                # Skip bold labels on computed rows — already classified as
                # row labels above.
                effective_bold = is_bold and _r not in _computed_rows_in_section
                if effective_bold:
                    # A single-bold row whose cell is wide-merged is a
                    # section-wide header, not a column label.
                    if _bold_per_row[_r] < 2 and _cs >= _WIDE_MERGE_THRESHOLD:
                        section_wide_labels[_r] = lbl_name
                        continue
                    if _c not in section_col_labels:
                        section_col_labels[_c] = lbl_name
                    if col_label_row is None or _r < col_label_row:
                        col_label_row = _r

            # --- Detect non-bold column headers in open/expandable tables ---
            # In sections like b1_list_of_sites and b1_list_of_subsidiaries,
            # the column headers are non-bold template_label cells that all
            # appear on the same row.  When a single row contains 2+ non-bold
            # label cells, those are column headers, not row labels.
            # Reclassify them: remove from row labels, add to column labels.
            if not section_col_labels:
                from collections import Counter

                non_bold_row_counts: Counter[int] = Counter()
                non_bold_cells_by_row: dict[int, list[tuple[int, int, str]]] = {}
                for _r, _c, lbl_name, is_bold, _cs in section_label_cells:
                    if (_r, _c) in title_positions:
                        continue
                    if not is_bold:
                        non_bold_row_counts[_r] += 1
                        non_bold_cells_by_row.setdefault(_r, []).append(
                            (_r, _c, lbl_name)
                        )

                for row_num, count in non_bold_row_counts.items():
                    if count >= 2:
                        # This row has multiple non-bold labels → column headers
                        for _r, _c, lbl_name in non_bold_cells_by_row[row_num]:
                            if _c not in section_col_labels:
                                section_col_labels[_c] = lbl_name
                            if col_label_row is None or _r < col_label_row:
                                col_label_row = _r
                        # Remove from row labels
                        section_row_labels.pop(row_num, None)

            # --- Index label reclassification ---------------------------------
            # A column that contains *both* a bold column header and non-bold
            # row labels below it is an **index column** (its header describes
            # what the row labels represent, e.g. "Land-use type", "Row ID",
            # "Pollutant").  These headers are not data column labels —
            # no XBRL data field sits in those columns — so they should be
            # stored as ``indexLabels`` instead of ``columnLabels``.
            #
            # Compute the set of columns that still contain row labels
            # (after the non-bold header reclassification above may have
            # removed some rows from section_row_labels).
            _row_label_cols: set[int] = {
                _row_label_col_map[r]
                for r in section_row_labels
                if r in _row_label_col_map
            }
            section_index_labels: dict[int, str] = {}  # col → template_label_name
            if _row_label_cols and section_col_labels:
                _idx_cols = set(section_col_labels.keys()) & _row_label_cols
                for _ic in _idx_cols:
                    section_index_labels[_ic] = section_col_labels.pop(_ic)

            section_id = hdr["name"].replace("template_", "")

            # Find the earliest additional_rows_warning in this section.
            # This marks where the repeatable table data area ends;
            # fields below this row are summary/total rows that should
            # not inherit table column labels.
            _arw_row: int | None = None
            for (_arw_sh, _arw_r), _arw_c in _additional_rows_warnings.items():
                if _arw_sh != hdr["sheet"]:
                    continue
                if hdr["row"] < _arw_r <= end_row:
                    if _arw_row is None or _arw_r < _arw_row:
                        _arw_row = _arw_r

            # Collect computed cell positions within this section for
            # marking fields as auto-calculated.
            _computed_in_section: set[tuple[int, int]] = set()
            for _cr_sh, _cr_r, _cr_c in _computed_cell_set:
                if _cr_sh == hdr["sheet"] and hdr["row"] <= _cr_r <= end_row:
                    _computed_in_section.add((_cr_r, _cr_c))

            excel_sections.append(
                {
                    "sectionId": section_id,
                    "templateName": hdr["name"],
                    "sheet": hdr["sheet"],
                    "headerRow": hdr["row"],
                    "headerCol": get_column_letter(hdr["col"]),
                    "endRow": end_row,
                    "colMin": col_min,
                    "colMax": col_max,
                    "applicability": hdr["applicability"],
                    "xbrlFields": xbrl_fields,
                    "yellowCellFields": yellow_fields_in_section,
                    "unitFields": unit_fields_in_section,
                    "fieldPositions": field_position_map,
                    "rowLabels": section_row_labels,
                    "columnLabels": section_col_labels,
                    "indexLabels": section_index_labels,
                    "colLabelRow": col_label_row,
                    "sectionWideLabels": section_wide_labels,
                    "additionalRowsWarningRow": _arw_row,
                    "computedPositions": _computed_in_section,
                }
            )

        # --- Inherit row labels for side-by-side sections --------------------
        # Some sections (e.g. C3 GHG reduction targets) are placed beside
        # another section (e.g. B3 GHG emissions) and share the same rows
        # but have no row labels of their own.  For these, merge row labels
        # from the sibling section that has them.  The sibling's labels are
        # added without overwriting any labels the section already owns.
        # Only inherit from a sibling that has strictly more row labels to
        # avoid bidirectional pollution (e.g. B3→C3 yes, C3→B3 no).
        for es in excel_sections:
            if not es["columnLabels"]:
                continue  # not a matrix section at all
            # Find a sibling: same sheet, same headerRow, different column range
            for sibling in excel_sections:
                if sibling is es:
                    continue
                if (
                    sibling["sheet"] == es["sheet"]
                    and sibling["headerRow"] == es["headerRow"]
                    and sibling["rowLabels"]
                    and len(sibling["rowLabels"]) > len(es["rowLabels"])
                ):
                    # Merge: sibling labels first, then overlay own labels
                    merged = dict(sibling["rowLabels"])
                    merged.update(es["rowLabels"])
                    es["rowLabels"] = merged
                    break

        # Build ``warningDefinitions``: a mapping from each warning key
        # (e.g. ``"NACE_warning"``) to the i18n translations extracted
        # from the Translations sheet.  The key is the suffix after
        # ``template_label_``.
        warning_definitions: dict[str, dict[str, str]] = {}
        for key, lang_map in translations.items():
            if "warning" in key:
                warning_definitions[key] = lang_map

        return {
            "labels": labels,
            "translations": translations,
            "metadata": metadata,
            "sectionApplicability": section_applicability,
            "sectionLabels": section_labels,
            "excelSections": excel_sections,
            "omittedDisclosures": omitted_disclosures,
            "enumLists": enum_lists,
            "xbrlToEnum": xbrl_to_enum,
            "definedNames": defined_names,
            "validationRules": validation_rules,
            "warningDefinitions": warning_definitions,
            "xbrlNameToTemplateLabel": _xbrl_name_to_template_label,
        }


STANDARD_LABEL_ROLE = "http://www.xbrl.org/2003/role/label"

# Mapping from XBRL dataType local names to survey-friendly inputType hints.
_DATA_TYPE_TO_INPUT_TYPE: dict[str, str] = {
    "booleanItemType": "boolean",
    "textBlockItemType": "textarea",
    "stringItemType": "text",
    "decimalItemType": "number",
    "integerItemType": "number",
    "monetaryItemType": "monetary",
    "percentItemType": "percent",
    "dateItemType": "date",
    "gYearItemType": "year",
    "enumerationItemType": "select",
    "enumerationSetItemType": "multiselect",
    # Physical-measurement types → number
    "massItemType": "number",
    "energyItemType": "number",
    "volumeItemType": "number",
    "areaItemType": "number",
    "ghgEmissionsItemType": "number",
    "ghgPerMonetaryItemType": "number",
}


def _input_type_for(concept) -> str:
    """Derive a survey-friendly inputType from a concept's dataType."""
    if concept.dataType:
        local = str(concept.dataType).rsplit(":", 1)[-1]
        if local in _DATA_TYPE_TO_INPUT_TYPE:
            return _DATA_TYPE_TO_INPUT_TYPE[local]
    # Fallback heuristics
    if concept.isBoolean:
        return "boolean"
    if concept.isMonetary:
        return "monetary"
    if concept.isNumeric:
        return "number"
    if concept.isTextblock:
        return "textarea"
    return "text"


def _concept_labels_i18n(concept) -> dict[str, str]:
    """Extract multilingual labels from a concept's _labels dict.

    Returns e.g. ``{"en": "English text", "de": "German text", …}``.
    """
    labels: dict[str, str] = {}
    for lang, roles in concept._labels.items():
        text = roles.get(STANDARD_LABEL_ROLE)
        if text:
            labels[lang] = text
    return labels


def _concept_to_field(
    concept,
    xbrl_to_enum: Optional[dict[str, list[str]]] = None,
    validation_rules: Optional[dict[str, dict]] = None,
) -> dict:
    """Build a survey-field descriptor from a taxonomy Concept.

    This is used both for facts already in the report (via ``fact_to_json_entry``)
    and for expected-but-empty fields listed in each section.

    If *validation_rules* is provided and contains an entry for this concept's
    local name, the field is annotated with ``validation`` info (required /
    conditional / informational, plus the condition field name if any).
    """
    local_name = (
        str(concept.qname).split(":", 1)[-1]
        if ":" in str(concept.qname)
        else str(concept.qname)
    )
    field: dict = {
        "qname": str(concept.qname),
        "label": concept.getStandardLabel(fallbackToQName=True),
        "labels": _concept_labels_i18n(concept),
        "dataType": str(concept.dataType) if concept.dataType else None,
        "periodType": concept.periodType,
        "inputType": _input_type_for(concept),
        "isRequired": concept.isReportable,
    }
    # Add validation info from the Excel IF-formula cells
    if validation_rules and local_name in validation_rules:
        field["validation"] = validation_rules[local_name]
    # Add enum options
    if xbrl_to_enum and local_name in xbrl_to_enum:
        enum_refs = xbrl_to_enum[local_name]
        field["options"] = enum_refs[0] if len(enum_refs) == 1 else enum_refs
    elif concept.isEnumerationSingle or concept.isEnumerationSet:
        # Try domain members from taxonomy
        domain = concept.getEEDomain()
        if domain:
            field["domainMembers"] = [
                {
                    "qname": str(m.qname),
                    "label": m.getStandardLabel(fallbackToQName=True),
                    "labels": _concept_labels_i18n(m),
                }
                for m in domain
            ]
    return field


def fact_to_json_entry(
    fact,
    xbrl_to_enum: Optional[dict[str, list[str]]] = None,
) -> dict:
    """Convert a single Fact to a JSON-serializable dict.

    If *xbrl_to_enum* is provided and the fact's concept name matches an entry,
    an ``"options"`` key is added referencing the ``enum_*`` list name(s) that
    supply the dropdown choices for this fact.
    """
    concept = fact.concept
    entry: dict = {
        "qname": str(concept.qname),
        "label": concept.getStandardLabel(fallbackToQName=True),
        "labels": _concept_labels_i18n(concept),
        "value": fact_value_to_json(fact.value),
        "dataType": str(concept.dataType) if concept.dataType else None,
        "periodType": concept.periodType,
        "inputType": _input_type_for(concept),
    }

    # Add formatted value if available
    try:
        formatted = fact.formattedValue
        if formatted and formatted != str(fact.value):
            entry["formattedValue"] = formatted
    except Exception:
        pass

    # Add aspects (period, units, dimensions, etc.)
    aspects = fact.aspects
    if aspects:
        serializable_aspects = {}
        for k, v in aspects.items():
            serializable_aspects[str(k)] = str(v)
        entry["aspects"] = serializable_aspects

    # Add taxonomy dimensions if present
    if fact.hasTaxonomyDimensions():
        dims = {}
        for dim_qname, member_qname in fact.getTaxonomyDimensions().items():
            dims[str(dim_qname)] = str(member_qname)
        entry["dimensions"] = dims

    # Add reference to the enum list(s) that provide dropdown options
    if xbrl_to_enum:
        # concept.qname is e.g. "vsme:SomeName"; the XBRL named range uses
        # just the local part "SomeName"
        local_name = (
            str(concept.qname).split(":", 1)[-1]
            if ":" in str(concept.qname)
            else str(concept.qname)
        )
        if local_name in xbrl_to_enum:
            enum_refs = xbrl_to_enum[local_name]
            entry["options"] = enum_refs[0] if len(enum_refs) == 1 else enum_refs

    # Add domain members from taxonomy for enum types without a mapped list
    if "options" not in entry and (
        concept.isEnumerationSingle or concept.isEnumerationSet
    ):
        domain = concept.getEEDomain()
        if domain:
            entry["domainMembers"] = [
                {
                    "qname": str(m.qname),
                    "label": m.getStandardLabel(fallbackToQName=True),
                    "labels": _concept_labels_i18n(m),
                }
                for m in domain
            ]

    return entry


def extract_facts_by_section(
    report,
    excel_sections: list[dict],
    flat: bool = False,
    xbrl_to_enum: Optional[dict[str, list[str]]] = None,
    validation_rules: Optional[dict[str, dict]] = None,
    translations: Optional[dict[str, dict[str, str]]] = None,
    xbrl_name_to_template_label: Optional[dict[str, str]] = None,
) -> list[dict]:
    """
    Build the sections list using the **Excel-native section structure**
    extracted from ``template_*`` defined names in the data sheets.

    Each ``excel_sections`` entry carries:

    * ``sectionId``    – e.g. ``"b3_total_energy_consumption_in_MWh"``
    * ``templateName`` – the full defined name
    * ``sheet``        – the data sheet name
    * ``headerRow``    – row number of the coloured header
    * ``applicability``– derived from the header formula
    * ``xbrlFields``   – list of XBRL defined-name strings in this section

    For each section we:

    1. Look up taxonomy concepts by matching the XBRL field name to the
       concept's local qname – this gives us multilingual labels, dataType,
       inputType, enum options, and validation info.
    2. Find reported facts whose concept matches any field in the section.
    """
    taxonomy = report._taxonomy

    # Build concept look-up: local name → Concept using taxonomy API
    concept_by_local: dict[str, object] = {}
    for concept in taxonomy._concepts.values():
        local = concept.qname.localName
        concept_by_local[local] = concept

    # Build fact look-up: concept local name → list[Fact]
    fact_by_local: dict[str, list] = {}
    for fact in report._facts:
        local = fact.concept.qname.localName
        fact_by_local.setdefault(local, []).append(fact)

    # Build mapping from base concept names to section xbrlField entries.
    # Excel defined names for dimensional data use suffixes like
    # ``_RenewableEnergyMember``, but the report facts use the base concept
    # name (e.g. ``EnergyConsumptionFromElectricity``).  We collect the base
    # names so we can also pull in dimensional facts.
    _MEMBER_SUFFIXES = re.compile(r"_[A-Z][A-Za-z]+Member$")
    _TABLE_AXIS_SUFFIXES = re.compile(r"(Table|Axis)$")
    _UNIT_SUFFIX = re.compile(r"_unit$")

    # Build a global set of base concept names that have at least one
    # compound (member-suffixed) variant across ALL sections.  This is
    # used to decide which concepts need dimension-aware filtering during
    # fact collection.  Concepts that only appear as bare names (no
    # compound variant anywhere) can safely accept all facts without
    # filtering — their dimensional facts won't collide with another
    # section's fields.
    _global_concepts_with_members: set[str] = set()
    for _es in excel_sections:
        for _fn in _es.get("xbrlFields", []):
            _m = _MEMBER_SUFFIXES.search(_fn)
            if _m:
                _global_concepts_with_members.add(_fn[: _m.start()])

    result: list[dict] = []
    for es in excel_sections:
        section_id = es["sectionId"]
        xbrl_fields = es.get("xbrlFields", [])
        yellow_cell_fields = es.get("yellowCellFields", {})
        field_positions = es.get("fieldPositions", {})
        section_row_labels = es.get("rowLabels", {})
        section_col_labels = es.get("columnLabels", {})
        unit_fields = es.get("unitFields", {})
        col_label_row: int | None = es.get("colLabelRow")
        section_wide_labels = es.get("sectionWideLabels", {})
        additional_rows_warning_row: int | None = es.get("additionalRowsWarningRow")
        computed_positions: set[tuple[int, int]] = es.get("computedPositions", set())

        # --- Expected fields from taxonomy -----------------------------------
        expected_by_local: dict[str, dict] = {}
        for field_name in xbrl_fields:
            if field_name in yellow_cell_fields:
                continue  # handled separately below
            concept = concept_by_local.get(field_name)
            if concept is None:
                continue
            if concept.isAbstract:
                continue
            field = _concept_to_field(concept, xbrl_to_enum, validation_rules)
            # Attach the template_label_* key and its i18n translations
            # so the survey can show the spreadsheet's question phrasing
            # alongside the XBRL taxonomy label.
            tl_key = (
                xbrl_name_to_template_label.get(field_name)
                if xbrl_name_to_template_label
                else None
            )
            field["labelKey"] = tl_key
            tl_labels: dict[str, str] = {}
            if tl_key and translations and tl_key in translations:
                tl_labels = dict(translations[tl_key])
            field["templateLabels"] = tl_labels
            expected_by_local[field_name] = field

        # --- Collect concept names reachable from this section ----------------
        # Direct names + base names derived from member-suffixed fields.
        # Also build a per-base-name set of expected member suffixes so that
        # dimensional-fact collection can filter out facts whose dimensions
        # don't correspond to defined names in *this* section.  This prevents
        # side-by-side sections (e.g. B3 GHG and C3 GHG targets) from
        # sharing the same facts despite covering different columns.
        base_names: set[str] = set()
        for fn in xbrl_fields:
            m = _MEMBER_SUFFIXES.search(fn)
            if m:
                base = fn[: m.start()]
                base_names.add(base)
            base2 = _TABLE_AXIS_SUFFIXES.sub("", fn)
            if base2 != fn:
                base_names.add(base2)
            base3 = _UNIT_SUFFIX.sub("", fn)
            if base3 != fn:
                base_names.add(base3)

        # --- Collect all reported facts for this section ----------------------
        seen_fact_ids: set[int] = set()
        # Map concept local name → list of raw Fact objects
        matched_facts: dict[str, list] = {}

        # For side-by-side sections that share the same base concept names
        # (e.g. B3 GHG and C3 GHG targets), facts must be filtered by
        # their XBRL dimension members to avoid duplication.
        #
        # Build a set of (base_concept, member_suffix) pairs expected by
        # this section:
        #   - Compound field ``Foo_BarMember`` → (Foo, BarMember)
        #   - Bare field ``Foo``               → (Foo, None)
        #
        # A fact with concept ``Foo`` and dimension member ``BarMember``
        # is only accepted if (Foo, BarMember) is in the expected set.
        # A fact with concept ``Foo`` and no dimensions is only accepted
        # if (Foo, None) is in the expected set.
        _expected_concept_members: set[tuple[str, str | None]] = set()
        # Track which base concepts have at least one compound (member)
        # variant *in this section* so we know where strict dimension
        # filtering is needed.
        _concepts_with_members: set[str] = set()
        # Track concepts where a _Total*Member compound exists — in XBRL
        # the "total" / default member is represented by the absence of
        # any dimension, so we need to also accept dimension-less facts.
        _concepts_with_total_member: set[str] = set()
        for fn in xbrl_fields:
            m = _MEMBER_SUFFIXES.search(fn)
            if m:
                base = fn[: m.start()]
                member = m.group()[1:]  # strip leading '_'
                _expected_concept_members.add((base, member))
                _concepts_with_members.add(base)
                if member.startswith("Total"):
                    _concepts_with_total_member.add(base)
            else:
                # bare name — only match facts without dimensions
                _expected_concept_members.add((fn, None))

        def _fact_accepted(base_concept: str, fact) -> bool:
            """Return True if *fact* belongs to this section based on its
            dimension members and the section's expected (concept, member)
            pairs.

            Dimension filtering is only applied when the concept has
            compound (member-suffixed) variants somewhere in the workbook
            (``_global_concepts_with_members``).  This prevents
            side-by-side sections (e.g. B3 GHG vs C3 GHG targets) from
            sharing facts, while leaving sections like B7 Waste (whose
            bare names have dimensional facts but no compound counterpart)
            unaffected.

            For concepts with a ``_Total*Member`` compound name (e.g.
            ``_TotalRenewableAndNonRenewableEnergyMember``), dimension-less
            facts are also accepted because in XBRL the "total" / default
            member is represented by the absence of any dimension.
            """
            # If no section in the workbook has compound variants for
            # this concept, accept all facts unconditionally.
            if base_concept not in _global_concepts_with_members:
                return True

            fact_members: set[str] = set()
            if fact.hasTaxonomyDimensions():
                for _dk, _dv in fact.getTaxonomyDimensions().items():
                    dv_str = str(_dv)
                    member = dv_str.split(":")[-1] if ":" in dv_str else dv_str
                    fact_members.add(member)
            if fact_members:
                return any(
                    (base_concept, m) in _expected_concept_members for m in fact_members
                )
            else:
                # Fact has no dimensions.  Accept if either:
                # - the bare name is an expected direct field, OR
                # - this section has a _Total*Member compound for the
                #   concept (the XBRL default member = no dimension).
                if (base_concept, None) in _expected_concept_members:
                    return True
                if base_concept in _concepts_with_total_member:
                    return True
                return False

        # Match facts by direct field name, filtered by dimension membership
        for field_name in xbrl_fields:
            facts = fact_by_local.get(field_name, [])
            for fact in facts:
                fid = id(fact)
                if fid in seen_fact_ids:
                    continue
                if _fact_accepted(field_name, fact):
                    seen_fact_ids.add(fid)
                    matched_facts.setdefault(field_name, []).append(fact)

        # Match dimensional facts by base concept name, filtered by
        # dimension membership.
        for base_name in base_names:
            facts = fact_by_local.get(base_name, [])
            for fact in facts:
                fid = id(fact)
                if fid in seen_fact_ids:
                    continue
                if _fact_accepted(base_name, fact):
                    seen_fact_ids.add(fid)
                    matched_facts.setdefault(base_name, []).append(fact)

        # --- Merge into a single `fields` array ------------------------------
        has_facts = False
        fields: list[dict] = []
        consumed_fact_keys: set[str] = set()

        if flat:
            # Flat mode: simple label→value dict (unchanged behaviour)
            flat_facts: dict[str, object] = {}
            for key, fact_list in matched_facts.items():
                for fact in fact_list:
                    has_facts = True
                    label = fact.concept.getStandardLabel(fallbackToQName=True)
                    value = fact_value_to_json(fact.value)
                    if label in flat_facts:
                        existing = flat_facts[label]
                        if isinstance(existing, list):
                            existing.append(value)
                        else:
                            flat_facts[label] = [existing, value]
                    else:
                        flat_facts[label] = value
            # In flat mode, keep the old structure
            fields_or_flat = flat_facts
        else:
            # Pre-compute sorted row-label rows for nearest-above lookup
            _sorted_rl_rows: list[int] = (
                sorted(section_row_labels.keys()) if section_row_labels else []
            )

            # Helper: build row/column label annotation for a field
            def _matrix_labels(
                field_name: str, *, is_yellow_cell: bool = False
            ) -> dict:
                """Return a dict with ``rowLabel`` and ``columnLabel``
                entries for *field_name* based on its cell position and
                the section's row/column label maps.

                Row labels: if the field's row has no exact match in
                ``section_row_labels``, fall back to the **nearest
                labelled row at or above** the field's row.  This handles
                the common pattern where the template_label cell is on a
                row above the XBRL data cell.

                Column labels are only applied to fields at or below the
                row where bold column headers appear (``col_label_row``).
                Fields above the table header area are question-answer
                pairs, not table cells, so they do not get column labels.

                Yellow-cell fields that resolve to a row label are
                standalone Q&A pairs (e.g. "Is the undertaking
                disclosing…?" at C29/G29) and should NOT receive a
                column label even if they sit below the column-header
                row, because they are not matrix data cells.
                """
                pos = field_positions.get(field_name)
                rl: dict | None = None
                cl: dict | None = None
                if pos:
                    f_row, f_col = pos
                    # Row label: exact match first, then nearest above
                    rl_name = section_row_labels.get(f_row)
                    if not rl_name and _sorted_rl_rows:
                        # Find the nearest row label at or above f_row
                        import bisect

                        idx = bisect.bisect_right(_sorted_rl_rows, f_row)
                        if idx > 0:
                            candidate_row = _sorted_rl_rows[idx - 1]
                            # Do not let the fallback cross the column-
                            # header row boundary: if the field sits in
                            # the table data area (at or below
                            # col_label_row) but the candidate row label
                            # is above col_label_row, the field is a
                            # table data cell without its own row index
                            # and should not inherit a Q&A row label
                            # from the section header area above.
                            if (
                                col_label_row is not None
                                and f_row >= col_label_row
                                and candidate_row < col_label_row
                            ):
                                pass  # do not inherit — leave rl_name as None
                            else:
                                rl_name = section_row_labels.get(candidate_row)
                    if rl_name:
                        rl_key = rl_name[len("template_label_") :]
                        rl_tl = {}
                        if translations and rl_key in translations:
                            rl_tl = dict(translations[rl_key])
                        rl = {
                            "key": rl_key,
                            "templateLabelKey": rl_name,
                            "templateLabels": rl_tl,
                        }
                    # Only assign column labels for fields in the table area
                    # (at or below the bold column header row, and above the
                    # additional_rows_warning boundary if one exists).
                    # Yellow-cell Q&A pairs (boolean trigger questions) that
                    # happen to sit below the header row should NOT get
                    # column labels — they are standalone questions, not
                    # table data cells.
                    # Fields at or below the additional_rows_warning row are
                    # summary/total rows outside the table's repeatable area.
                    suppress_col = is_yellow_cell and rl is not None
                    below_warning = (
                        additional_rows_warning_row is not None
                        and f_row > additional_rows_warning_row
                    )
                    if (
                        col_label_row is not None
                        and f_row >= col_label_row
                        and not suppress_col
                        and not below_warning
                    ):
                        cl_name = section_col_labels.get(f_col)
                        if cl_name:
                            cl_key = cl_name[len("template_label_") :]
                            cl_tl = {}
                            if translations and cl_key in translations:
                                cl_tl = dict(translations[cl_key])
                            cl = {
                                "key": cl_key,
                                "templateLabelKey": cl_name,
                                "templateLabels": cl_tl,
                            }
                return {"rowLabel": rl, "columnLabel": cl}

            # For each field in section order, merge with matching facts.
            # Fields can be either taxonomy-backed XBRL fields or yellow-cell
            # helper fields.  We iterate xbrl_fields to preserve row order.
            _emitted_unit_cells: set[str] = set()
            for field_name in xbrl_fields:
                # --- Yellow-cell helper field ---------------------------------
                if field_name in yellow_cell_fields:
                    yf = yellow_cell_fields[field_name]
                    lk = yf.get("labelKey")
                    tl = {}
                    if lk and translations and lk in translations:
                        tl = dict(translations[lk])
                    entry = {
                        "fieldId": yf["fieldId"],
                        "source": "yellowCell",
                        "labelKey": lk,
                        "templateLabelKey": yf.get("templateLabelKey"),
                        "label": yf["label"],
                        "labels": yf.get("labels", {}),
                        "templateLabels": tl,
                        "dataType": yf.get("dataType"),
                        "inputType": yf.get("inputType", "text"),
                        "isRequired": False,
                        "isReportable": False,
                        "value": yf.get("value"),
                        "cellRef": yf.get("cellRef"),
                    }
                    if yf.get("options"):
                        entry["options"] = yf["options"]
                    if yf.get("isComputed"):
                        entry["isComputed"] = True
                    if yf.get("value") is not None:
                        has_facts = True
                    entry.update(_matrix_labels(field_name, is_yellow_cell=True))
                    fields.append(entry)
                    continue

                # --- Unit-selection field -------------------------------------
                if field_name in unit_fields:
                    uf = unit_fields[field_name]
                    cell_ref = uf.get("cellRef")

                    # De-duplicate: when multiple XBRL _unit fields
                    # reference the same physical cell (e.g. one unit
                    # dropdown shared by several data rows), emit only
                    # a single consolidated field and record all the
                    # constituent field IDs in ``unitFieldIds``.
                    if cell_ref and cell_ref in _emitted_unit_cells:
                        # Already emitted — append this field ID to
                        # the existing entry's unitFieldIds list.
                        for prev in fields:
                            if (
                                prev.get("source") == "unitSelection"
                                and prev.get("cellRef") == cell_ref
                            ):
                                prev.setdefault("unitFieldIds", [prev["fieldId"]])
                                if uf["fieldId"] not in prev["unitFieldIds"]:
                                    prev["unitFieldIds"].append(uf["fieldId"])
                                break
                        continue

                    lk = uf.get("labelKey")
                    tl = uf.get("labels", {})
                    # If no direct template label mapping, try the section-
                    # wide label nearest above the unit cell.  Section-wide
                    # labels are bold cells on rows with only one bold label
                    # (e.g. "Please select the unit used for reporting…").
                    if not lk and section_wide_labels:
                        pos = field_positions.get(field_name)
                        if pos:
                            f_row = pos[0]
                            import bisect as _bisect

                            sw_rows = sorted(section_wide_labels.keys())
                            idx = _bisect.bisect_right(sw_rows, f_row)
                            if idx > 0:
                                sw_name = section_wide_labels[sw_rows[idx - 1]]
                                lk = sw_name[len("template_label_") :]
                                if translations and lk in translations:
                                    tl = dict(translations[lk])
                    # Last fallback: column label at the field's column
                    if not lk:
                        pos = field_positions.get(field_name)
                        if pos:
                            f_col = pos[1]
                            cl_name = section_col_labels.get(f_col)
                            if cl_name:
                                lk = cl_name[len("template_label_") :]
                                if translations and lk in translations:
                                    tl = dict(translations[lk])
                    entry = {
                        "fieldId": uf["fieldId"],
                        "source": "unitSelection",
                        "labelKey": lk,
                        "templateLabelKey": (f"template_label_{lk}" if lk else None),
                        "label": tl.get("en", uf["fieldId"]),
                        "labels": tl,
                        "templateLabels": tl,
                        "dataType": "enumeration",
                        "inputType": "dropdown",
                        "isRequired": False,
                        "isReportable": False,
                        "value": uf.get("value"),
                        "cellRef": uf.get("cellRef"),
                    }
                    if uf.get("options"):
                        entry["options"] = uf["options"]
                    if uf.get("validation"):
                        entry["validation"] = uf["validation"]
                        entry["isRequired"] = True
                    if uf.get("value") is not None:
                        has_facts = True
                    # Unit-selection fields do NOT use matrix labels.
                    # Their label comes from the section-wide label
                    # resolution above.  They are standalone dropdowns
                    # that sit between the Q&A area and the table area.
                    entry["rowLabel"] = None
                    entry["columnLabel"] = None
                    fields.append(entry)
                    if cell_ref:
                        _emitted_unit_cells.add(cell_ref)
                    continue

                # --- Taxonomy-backed XBRL field -------------------------------
                schema = expected_by_local.get(field_name)
                if schema is None:
                    continue  # abstract concept or not found in taxonomy

                fact_list = matched_facts.get(field_name, [])
                consumed_fact_keys.add(field_name)

                if not fact_list:
                    # No reported value — emit schema with value: null
                    entry = dict(schema)
                    entry["value"] = None
                    ml = _matrix_labels(field_name)
                    # Mark computed fields (value cell has auto-calc fill)
                    pos = field_positions.get(field_name)
                    if pos and (pos[0], pos[1]) in computed_positions:
                        entry["isComputed"] = True
                    entry.update(ml)
                    fields.append(entry)
                else:
                    # One or more reported facts for this field
                    ml = _matrix_labels(field_name)
                    # Check if the field's value cell is auto-calculated
                    pos = field_positions.get(field_name)
                    _field_is_computed = bool(
                        pos and (pos[0], pos[1]) in computed_positions
                    )
                    for fact in fact_list:
                        has_facts = True
                        entry = dict(schema)
                        entry["value"] = fact_value_to_json(fact.value)
                        # Add formatted value if available
                        try:
                            formatted = fact.formattedValue
                            if formatted and formatted != str(fact.value):
                                entry["formattedValue"] = formatted
                        except Exception:
                            pass
                        # Add aspects
                        aspects = fact.aspects
                        if aspects:
                            entry["aspects"] = {
                                str(k): str(v) for k, v in aspects.items()
                            }
                        # Add taxonomy dimensions if present
                        if fact.hasTaxonomyDimensions():
                            entry["dimensions"] = {
                                str(dk): str(dv)
                                for dk, dv in fact.getTaxonomyDimensions().items()
                            }
                        entry.update(ml)
                        if _field_is_computed:
                            entry["isComputed"] = True
                        fields.append(entry)

            # Append any extra facts not covered by expectedFields (e.g.
            # dimensional facts matched via base-name derivation)
            for key, fact_list in matched_facts.items():
                if key in consumed_fact_keys:
                    continue
                for fact in fact_list:
                    has_facts = True
                    fact_entry = fact_to_json_entry(fact, xbrl_to_enum=xbrl_to_enum)
                    fact_entry.setdefault("value", None)
                    # Attach template label linkage for extra dimensional facts
                    local = fact.concept.qname.localName
                    tl_key = (
                        xbrl_name_to_template_label.get(local)
                        if xbrl_name_to_template_label
                        else None
                    )
                    fact_entry.setdefault("labelKey", tl_key)
                    tl_labels: dict[str, str] = {}
                    if tl_key and translations and tl_key in translations:
                        tl_labels = dict(translations[tl_key])
                    fact_entry.setdefault("templateLabels", tl_labels)
                    # Try to reconstruct the compound defined-name from the
                    # base concept name + dimension member suffix so that we
                    # can look up the cell position for matrix labels.
                    ml = {"rowLabel": None, "columnLabel": None}
                    if fact.hasTaxonomyDimensions():
                        dims = fact.getTaxonomyDimensions()
                        for _dk, _dv in dims.items():
                            member = (
                                str(_dv).split(":")[-1] if ":" in str(_dv) else str(_dv)
                            )
                            compound = f"{local}_{member}"
                            if compound in field_positions:
                                ml = _matrix_labels(compound)
                                break
                    if ml["rowLabel"] is None and ml["columnLabel"] is None:
                        # Fallback: try the base local name directly
                        if local in field_positions:
                            ml = _matrix_labels(local)
                    if ml["rowLabel"] is None and ml["columnLabel"] is None:
                        # Fallback for aggregate/total/default facts without
                        # explicit dimension members: search for any compound
                        # defined name starting with the base concept name.
                        # Priority order:
                        #   1. ``_CurrentlyStatedMember`` (the "default"
                        #      reporting column in GHG sections)
                        #   2. Any compound name containing "Total" in the
                        #      member suffix (e.g. energy breakdown totals)
                        #   3. Any remaining compound name (last resort)
                        currently_stated = f"{local}_CurrentlyStatedMember"
                        if currently_stated in field_positions:
                            ml = _matrix_labels(currently_stated)
                        else:
                            # Try "Total" names first
                            prefix = local + "_"
                            total_match = None
                            any_match = None
                            for fp_name in field_positions:
                                if fp_name.startswith(prefix):
                                    if any_match is None:
                                        any_match = fp_name
                                    if "Total" in fp_name[len(prefix) :]:
                                        total_match = fp_name
                                        break
                            best = total_match or any_match
                            if best:
                                ml = _matrix_labels(best)
                    fact_entry.update(ml)
                    fields.append(fact_entry)

            fields_or_flat = fields

        # Build the section entry
        # Prepare section-level row/column label descriptors with i18n.
        # Keys are the full ``template_label_*`` name; values include
        # the short key and all available translations.
        def _label_descriptor(full_name: str) -> dict:
            short_key = full_name[len("template_label_") :]
            tl_map: dict[str, str] = {}
            if translations and short_key in translations:
                tl_map = dict(translations[short_key])
            return {
                "key": short_key,
                "templateLabelKey": full_name,
                "templateLabels": tl_map,
            }

        row_label_descriptors: dict[str, dict] = {
            str(row): _label_descriptor(name)
            for row, name in section_row_labels.items()
        }
        col_label_descriptors: dict[str, dict] = {
            get_column_letter(col): _label_descriptor(name)
            for col, name in section_col_labels.items()
        }
        section_index_labels_raw = es.get("indexLabels", {})
        index_label_descriptors: dict[str, dict] = {
            get_column_letter(col): _label_descriptor(name)
            for col, name in section_index_labels_raw.items()
        }

        section_entry: dict = {
            "sectionId": section_id,
            "templateName": es["templateName"],
            "sheet": es["sheet"],
            "headerRow": es["headerRow"],
            "applicability": es["applicability"],
            "hasFacts": has_facts,
            "hasAdditionalRowsWarning": es.get("additionalRowsWarningRow") is not None,
            "rowLabels": row_label_descriptors if row_label_descriptors else None,
            "columnLabels": col_label_descriptors if col_label_descriptors else None,
            "indexLabels": index_label_descriptors if index_label_descriptors else None,
            "fields": fields_or_flat,
        }

        result.append(section_entry)

    return result


def main() -> None:
    parser = createArgParser()
    args = parseArgs(parser)

    resultsBuilder = ConversionResultsBuilder(consoleOutput=True)
    with resultsBuilder.processingContext("mireport Excel to JSON export") as pc:
        pc.mark("Loading taxonomy metadata")
        mireport.loadTaxonomyJSON()

        pc.mark(
            "Extracting data from Excel",
            additionalInfo=f"Using file: {args.excel_file}",
        )
        excel = JsonExcelProcessor(
            args.excel_file,
            resultsBuilder,
            VSME_DEFAULTS,
            outputLocale=args.output_locale,
        )
        report = excel.populateReport()

        pc.mark("Extracting template metadata & conditional questions")
        template_data = excel.get_template_data()

        pc.mark("Generating JSON output")
        sections_data = extract_facts_by_section(
            report,
            excel_sections=template_data["excelSections"],
            flat=args.flat,
            xbrl_to_enum=template_data["xbrlToEnum"],
            validation_rules=template_data["validationRules"],
            translations=template_data["translations"],
            xbrl_name_to_template_label=template_data["xbrlNameToTemplateLabel"],
        )

        # Build the output structure
        output = {
            "reportTitle": report._reportTitle,
            "entityName": report._entityName,
            "factCount": report.factCount,
            "templateMetadata": template_data["metadata"],
            "templateLabels": template_data["labels"],
            "sectionLabels": template_data["sectionLabels"],
            "translations": template_data["translations"],
            "sectionApplicability": template_data["sectionApplicability"],
            "omittedDisclosures": template_data["omittedDisclosures"],
            "sections": sections_data,
            "enumLists": template_data["enumLists"],
            "xbrlToEnum": template_data["xbrlToEnum"],
            "validationRules": template_data["validationRules"],
            "warningDefinitions": template_data["warningDefinitions"],
            "definedNames": template_data["definedNames"],
        }

        # Resolve output path
        output_path = args.output_path
        if output_path.is_dir() or (not output_path.suffix):
            output_path.mkdir(parents=True, exist_ok=True)
            output_path = output_path / "report-facts.json"
        else:
            output_path.parent.mkdir(parents=True, exist_ok=True)

        if output_path.exists() and not args.force:
            print(f"⚠️  Warning: Overwriting existing file: {output_path}")

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        pc.mark(
            "Done", additionalInfo=f"Wrote {report.factCount} facts to {output_path}"
        )

    result = resultsBuilder.build()
    if result.hasMessages(userOnly=True):
        print()
        messages = result.userMessages
        print(f"Messages ({len(messages)}):")
        for msg in messages:
            print(f"\t{msg}")


if __name__ == "__main__":
    rich.traceback.install(show_locals=False)
    logging.basicConfig(
        format="%(message)s",
        datefmt="[%Y-%m-%d %H:%M:%S]",
        handlers=[RichHandler(rich_tracebacks=True)],
    )
    logging.captureWarnings(True)
    main()
